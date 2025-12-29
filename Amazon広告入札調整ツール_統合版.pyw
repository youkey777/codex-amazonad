# -*- coding: utf-8 -*-
"""
Amazon広告 入札単価自動調整ツール v14
GUI版 - 統合形式対応版
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from tkinterdnd2 import DND_FILES, TkinterDnD
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl import Workbook
import os
import threading
from datetime import datetime


# ==================== 列インデックス定義（固定）====================
# Amazonレポートの列は固定形式なので、動的検出は不要

# SPキャンペーンシート（旧形式 - 54列）
SP_COL = {
    'KEYWORD_ID': 7,      # H列
    'TARGETING_ID': 8,    # I列
    'PORTFOLIO': 13,      # N列
    'BID': 27,            # AB列
    'KEYWORD_TEXT': 28,   # AC列（キーワードテキスト）
    'PRODUCT_TARGETING': 35,  # AJ列（商品ターゲティング式）
    'CLICKS': 41,         # AP列
    'SPEND': 43,          # AR列
    'SALES': 44,          # AS列
    'ORDERS': 45,         # AT列
    'ACOS': 48,           # AW列
    'TARGET_ACOS': 53,    # BB列
}

# SPキャンペーンシート（統合形式 - 55列）
# 新形式では40列目に「セグメント名」が追加され、41列目以降が+1シフト
UNIFIED_SP_COL = {
    'CAMPAIGN_ID': 3,     # D列（キャンペーンID）
    'KEYWORD_ID': 7,      # H列
    'TARGETING_ID': 8,    # I列
    'PORTFOLIO': 13,      # N列
    'TARGETING_TYPE': 16, # Q列（ターゲティングの種類: マニュアル/オート）
    'BID': 27,            # AB列
    'KEYWORD_TEXT': 28,   # AC列（キーワードテキスト）
    'MATCH_TYPE': 31,     # AF列（マッチタイプ: 完全一致/フレーズ/部分一致）
    'PRODUCT_TARGETING': 35,  # AJ列（商品ターゲティング式）
    'CLICKS': 42,         # AQ列（旧形式から+1シフト）
    'SPEND': 44,          # AS列（旧形式から+1シフト）
    'SALES': 45,          # AT列（旧形式から+1シフト）
    'ORDERS': 46,         # AU列（旧形式から+1シフト）
    'ACOS': 49,           # AX列（旧形式から+1シフト）
    'TARGET_ACOS': 54,    # BC列（旧形式から+1シフト）
}

# SP検索ワードシート（許容ACOSは検索ワードシートには存在しない）
SP_SEARCH_COL = {
    'CAMPAIGN_ID': 1,     # B列（キャンペーンID）
    'AD_GROUP_ID': 2,     # C列（広告グループID）
    'KEYWORD_ID': 3,      # D列
    'TARGETING_ID': 4,    # E列
    'CAMPAIGN_NAME': 5,   # F列（キャンペーン名）
    'AD_GROUP_NAME': 6,   # G列（広告グループ名）
    'PORTFOLIO': 7,       # H列
    'SEARCH_TERM': 15,    # P列
    'CLICKS': 17,         # R列
    'SPEND': 19,          # T列
    'SALES': 20,          # U列
    'ORDERS': 21,         # V列
    # TARGET_ACOSは検索ワードシートには存在しない - キャンペーンシートから取得する
}

# SBキャンペーンシート（旧形式 - 52列）
SB_COL = {
    'CAMPAIGN_ID': 3,     # D列（キャンペーンID）
    'KEYWORD_ID': 7,      # H列
    'TARGETING_ID': 8,    # I列
    'PORTFOLIO': 11,      # L列（ポートフォリオ名）
    'BID': 21,            # V列（入札額）
    'KEYWORD_TEXT': 22,   # W列（キーワードテキスト）
    'PRODUCT_TARGETING': 24,  # Y列（商品ターゲティング式）
    'CLICKS': 41,         # AP列
    'SPEND': 43,          # AR列（支出）
    'SALES': 44,          # AS列
    'ORDERS': 45,         # AT列
    'ACOS': 48,           # AW列
    'TARGET_ACOS': 51,    # AZ列（許容ACOS）
}

# SBキャンペーンシート（統合形式も旧形式と同じ52列構造）
# 統合形式にはターゲティングの種類とマッチタイプが含まれる
UNIFIED_SB_COL = SB_COL.copy()
UNIFIED_SB_COL.update({
    'TARGETING_TYPE': 16,  # Q列（ターゲティングの種類: マニュアル/オート）※要確認
    'MATCH_TYPE': 31,      # AF列（マッチタイプ: 完全一致/フレーズ/部分一致）※要確認
})

# SB検索ワードシート（許容ACOSは検索ワードシートには存在しない）
SB_SEARCH_COL = {
    'CAMPAIGN_ID': 1,     # B列（キャンペーンID）
    'AD_GROUP_ID': 2,     # C列（広告グループID）
    'KEYWORD_ID': 3,      # D列
    'TARGETING_ID': 4,    # E列
    'CAMPAIGN_NAME': 5,   # F列（キャンペーン名）
    'AD_GROUP_NAME': 6,   # G列（広告グループ名）
    'PORTFOLIO': 5,       # F列（SB検索ワードシートにはポートフォリオ列がないためキャンペーン名を使用）
    'SEARCH_TERM': 13,    # N列（カスタマー検索用語）
    'CLICKS': 15,         # P列
    'SPEND': 17,          # R列（支出）
    'SALES': 18,          # S列
    'ORDERS': 19,         # T列
    # TARGET_ACOSは検索ワードシートには存在しない - キャンペーンシートから取得する
}

# SDキャンペーンシート（旧形式 - 48列）
SD_COL = {
    'TARGETING_ID': 7,    # H列（ターゲティングID）
    'BID': 25,            # Z列（入札額）
    'CLICKS': 31,         # AF列（クリック数）
    'SPEND': 33,          # AH列（支出）
    'SALES': 34,          # AI列（売上）
    'ORDERS': 35,         # AJ列（注文数）
    'ACOS': 38,           # AM列（ACOS）
    'TARGET_ACOS': 47,    # AV列（許容ACOS）
}

# SDキャンペーンシート（統合形式も旧形式と同じ48列構造）
UNIFIED_SD_COL = SD_COL  # 新旧で列構造が同じため、SD_COLを使用


# ==================== スタイル定義 ====================
class AppStyle:
    BG_PRIMARY = "#FFFFFF"
    BG_SECONDARY = "#F5F7FA"
    SURFACE = "#FFFFFF"
    SURFACE_VARIANT = "#F8FAFC"
    PRIMARY = "#6366F1"
    PRIMARY_LIGHT = "#818CF8"
    SUCCESS = "#10B981"
    SUCCESS_LIGHT = "#D1FAE5"
    ERROR = "#EF4444"
    ERROR_LIGHT = "#FEE2E2"
    SP_COLOR = "#EC4899"
    SP_BG = "#FDF2F8"
    SB_COLOR = "#14B8A6"
    SB_BG = "#F0FDFA"
    SD_COLOR = "#8B5CF6"
    SD_BG = "#F5F3FF"
    TEXT_PRIMARY = "#1E293B"
    TEXT_SECONDARY = "#64748B"
    TEXT_TERTIARY = "#94A3B8"
    TEXT_ON_PRIMARY = "#FFFFFF"
    BORDER = "#E2E8F0"

    FONT_FAMILY = "Meiryo"
    FONT_DISPLAY = (FONT_FAMILY, 20, "bold")
    FONT_HEADLINE = (FONT_FAMILY, 16, "bold")
    FONT_TITLE = (FONT_FAMILY, 14, "bold")
    FONT_BODY = (FONT_FAMILY, 11)
    FONT_BODY_BOLD = (FONT_FAMILY, 11, "bold")
    FONT_CAPTION = (FONT_FAMILY, 10)
    FONT_SMALL = (FONT_FAMILY, 9)
    FONT_EMOJI = ("Segoe UI Emoji", 24)
    FONT_EMOJI_LARGE = ("Segoe UI Emoji", 40)
    FONT_EMOJI_SMALL = ("Segoe UI Emoji", 14)


# ==================== UIコンポーネント ====================
class ModernButton(tk.Canvas):
    def __init__(self, parent, text, command, bg_color, fg_color="#FFFFFF",
                 hover_color=None, width=180, height=48, **kwargs):
        super().__init__(parent, width=width, height=height,
                        bg=parent.cget('bg'), highlightthickness=0, **kwargs)
        self.command = command
        self.bg_color = bg_color
        self.hover_color = hover_color or self._lighten_color(bg_color)
        self.fg_color = fg_color
        self.text = text
        self.width = width
        self.height = height
        self.pressed = False
        self._draw_button(self.bg_color)
        self.bind("<Enter>", self._on_enter)
        self.bind("<Leave>", self._on_leave)
        self.bind("<Button-1>", self._on_press)
        self.bind("<ButtonRelease-1>", self._on_release)

    def _lighten_color(self, color):
        r, g, b = int(color[1:3], 16), int(color[3:5], 16), int(color[5:7], 16)
        return f"#{min(255, int(r*1.15)):02x}{min(255, int(g*1.15)):02x}{min(255, int(b*1.15)):02x}"

    def _darken_color(self, color):
        r, g, b = int(color[1:3], 16), int(color[3:5], 16), int(color[5:7], 16)
        return f"#{int(r*0.85):02x}{int(g*0.85):02x}{int(b*0.85):02x}"

    def _draw_button(self, color):
        self.delete("all")
        r, w, h = 12, self.width, self.height
        self.create_arc(0, 0, r*2, r*2, start=90, extent=90, fill=color, outline=color)
        self.create_arc(w-r*2, 0, w, r*2, start=0, extent=90, fill=color, outline=color)
        self.create_arc(0, h-r*2, r*2, h, start=180, extent=90, fill=color, outline=color)
        self.create_arc(w-r*2, h-r*2, w, h, start=270, extent=90, fill=color, outline=color)
        self.create_rectangle(r, 0, w-r, h, fill=color, outline=color)
        self.create_rectangle(0, r, w, h-r, fill=color, outline=color)
        self.create_text(w//2, h//2, text=self.text, fill=self.fg_color, font=AppStyle.FONT_BODY_BOLD)

    def _on_enter(self, e): self._draw_button(self.hover_color); self.config(cursor="hand2")
    def _on_leave(self, e): self._draw_button(self.bg_color)
    def _on_press(self, e): self.pressed = True; self._draw_button(self._darken_color(self.bg_color))
    def _on_release(self, e):
        if self.pressed:
            self.pressed = False
            self._draw_button(self.hover_color)
            if self.command: self.command()

    def set_state(self, state):
        if state == "disabled":
            self.bg_color = AppStyle.TEXT_TERTIARY
            self._draw_button(self.bg_color)
            for event in ["<Enter>", "<Leave>", "<Button-1>", "<ButtonRelease-1>"]:
                self.unbind(event)
        else:
            self.bg_color = AppStyle.PRIMARY
            self._draw_button(self.bg_color)
            self.bind("<Enter>", self._on_enter)
            self.bind("<Leave>", self._on_leave)
            self.bind("<Button-1>", self._on_press)
            self.bind("<ButtonRelease-1>", self._on_release)


# ==================== メインアプリ ====================
class AmazonBidAdjusterApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Amazon広告 入札調整ツール (統合版)")
        self.root.geometry("920x850")
        self.root.resizable(True, True)
        self.root.configure(bg=AppStyle.BG_SECONDARY)

        try:
            from ctypes import windll
            windll.shcore.SetProcessDpiAwareness(1)
        except: pass

        self.file_path = tk.StringVar()
        self.file_type = tk.StringVar(value="")
        self.full_path = None
        self.file_loaded = False
        self.file_format = None  # 新規: "unified" or "single_XX"
        self.ad_types_in_file = None  # 新規: 統合形式の場合の広告タイプリスト

        self.params = {
            'click_threshold_low': tk.IntVar(value=10),
            'click_threshold_mid': tk.IntVar(value=20),
            'click_threshold_high': tk.IntVar(value=30),
            'reflect_rate_low': tk.IntVar(value=30),
            'reflect_rate_mid': tk.IntVar(value=60),
            'reflect_rate_high': tk.IntVar(value=100),
            'reduce_rate_mid': tk.IntVar(value=25),
            'reduce_rate_high': tk.IntVar(value=50),
            'max_change': tk.IntVar(value=30),
            'min_bid': tk.IntVar(value=10),
            'top_n': tk.IntVar(value=30),
            'acos_protect_min': tk.IntVar(value=30),
            'acos_protect_max': tk.IntVar(value=35),
            'new_kw_cpc_add_3plus': tk.IntVar(value=30),
            'new_kw_cpc_add_2': tk.IntVar(value=15),
            'new_kw_order1_max': tk.IntVar(value=60),
        }

        self.create_ui()

    def create_ui(self):
        container = tk.Frame(self.root, bg=AppStyle.BG_SECONDARY)
        container.pack(fill=tk.BOTH, expand=True)

        canvas = tk.Canvas(container, bg=AppStyle.BG_SECONDARY, highlightthickness=0)
        scrollbar = ttk.Scrollbar(container, orient="vertical", command=canvas.yview)

        self.main_frame = tk.Frame(canvas, bg=AppStyle.BG_SECONDARY)
        canvas.create_window((0, 0), window=self.main_frame, anchor="nw")
        self.main_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.bind_all("<MouseWheel>", lambda e: canvas.yview_scroll(int(-1*(e.delta/120)), "units"))

        content = tk.Frame(self.main_frame, bg=AppStyle.BG_SECONDARY)
        content.pack(fill=tk.BOTH, expand=True, padx=32, pady=24)

        self._create_header(content)
        self._create_drop_zone(content)
        # ファイル情報コンテナ（調整パラメータの上に配置）
        self.file_info_container = tk.Frame(content, bg=AppStyle.BG_SECONDARY)
        # 初期状態では非表示
        self._create_params_section(content)
        self._create_action_buttons(content)
        self._create_log_area(content)

        self.log("Excelファイルをドロップしてください")

    def _create_header(self, parent):
        header = tk.Frame(parent, bg=AppStyle.BG_SECONDARY)
        header.pack(fill=tk.X, pady=(0, 24))

        title_frame = tk.Frame(header, bg=AppStyle.BG_SECONDARY)
        title_frame.pack(anchor="w")

        tk.Label(title_frame, text="Amazon広告 入札調整ツール", font=AppStyle.FONT_DISPLAY, fg=AppStyle.TEXT_PRIMARY, bg=AppStyle.BG_SECONDARY).pack(side=tk.LEFT)

    def _create_drop_zone(self, parent):
        self.drop_outer = tk.Frame(parent, bg=AppStyle.BORDER, padx=2, pady=2)
        self.drop_outer.pack(fill=tk.X, pady=(0, 16))

        self.drop_zone = tk.Frame(self.drop_outer, bg=AppStyle.SURFACE, padx=32, pady=32)
        self.drop_zone.pack(fill=tk.X)

        self.drop_inner = tk.Frame(self.drop_zone, bg=AppStyle.SURFACE)
        self.drop_inner.pack(fill=tk.X, padx=2, pady=2)

        self.drop_icon = tk.Label(self.drop_inner, text="📁", font=AppStyle.FONT_EMOJI, bg=AppStyle.SURFACE)
        self.drop_icon.pack(pady=(8, 12))

        self.drop_text = tk.Label(self.drop_inner, text="Excelファイルをここにドロップ", font=AppStyle.FONT_TITLE, fg=AppStyle.TEXT_PRIMARY, bg=AppStyle.SURFACE)
        self.drop_text.pack()

        self.drop_hint = tk.Label(self.drop_inner, text="またはクリックして選択", font=AppStyle.FONT_CAPTION, fg=AppStyle.TEXT_TERTIARY, bg=AppStyle.SURFACE)
        self.drop_hint.pack(pady=(6, 8))

        for w in [self.drop_zone, self.drop_inner, self.drop_icon, self.drop_text, self.drop_hint]:
            w.drop_target_register(DND_FILES)
            w.dnd_bind('<<Drop>>', self.on_drop)
            w.bind('<Button-1>', self.on_click_select)

    def _create_params_section(self, parent):
        # パラメータセクション全体を包むフレーム
        self.params_section_frame = tk.Frame(parent, bg=AppStyle.BG_SECONDARY)
        self.params_section_frame.pack(fill=tk.X)

        section_header = tk.Frame(self.params_section_frame, bg=AppStyle.BG_SECONDARY)
        section_header.pack(fill=tk.X, pady=(8, 12))
        tk.Label(section_header, text="調整パラメータ", font=AppStyle.FONT_HEADLINE, fg=AppStyle.TEXT_PRIMARY, bg=AppStyle.BG_SECONDARY).pack(side=tk.LEFT)

        cards = tk.Frame(self.params_section_frame, bg=AppStyle.BG_SECONDARY)
        cards.pack(fill=tk.X, pady=(0, 16))

        left = tk.Frame(cards, bg=AppStyle.BG_SECONDARY)
        left.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 8))

        right = tk.Frame(cards, bg=AppStyle.BG_SECONDARY)
        right.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(8, 0))

        self._create_param_card(left, "クリック閾値", AppStyle.PRIMARY, [
            ("低", "click_threshold_low", ""),
            ("中", "click_threshold_mid", ""),
            ("高", "click_threshold_high", ""),
        ])

        self._create_param_card(left, "制限値", AppStyle.SD_COLOR, [
            ("最大変動", "max_change", "円"),
            ("最小入札", "min_bid", "円"),
            ("上位N件", "top_n", ""),
            ("ACOS下限", "acos_protect_min", "%"),
            ("ACOS上限", "acos_protect_max", "%"),
        ])

        self._create_param_card(right, "反映率", AppStyle.SB_COLOR, [
            ("低", "reflect_rate_low", "%"),
            ("中", "reflect_rate_mid", "%"),
            ("高", "reflect_rate_high", "%"),
        ])

        self._create_param_card(right, "削減率", AppStyle.SP_COLOR, [
            ("中", "reduce_rate_mid", "%"),
            ("高", "reduce_rate_high", "%"),
        ])

        self._create_param_card(right, "新規KW候補", AppStyle.PRIMARY, [
            ("3件以上+", "new_kw_cpc_add_3plus", "円"),
            ("1-2件+", "new_kw_cpc_add_2", "円"),
            ("1件上限", "new_kw_order1_max", "円"),
        ])

    def _create_param_card(self, parent, title, accent, params):
        outer = tk.Frame(parent, bg=AppStyle.BORDER, padx=1, pady=1)
        outer.pack(fill=tk.X, pady=(0, 12))

        card = tk.Frame(outer, bg=AppStyle.SURFACE, padx=16, pady=14)
        card.pack(fill=tk.X)

        header = tk.Frame(card, bg=AppStyle.SURFACE)
        header.pack(fill=tk.X, pady=(0, 12))

        accent_line = tk.Frame(header, bg=accent, width=3, height=20)
        accent_line.pack(side=tk.LEFT, padx=(0, 10))
        accent_line.pack_propagate(False)

        tk.Label(header, text=title, font=AppStyle.FONT_BODY_BOLD, fg=AppStyle.TEXT_PRIMARY, bg=AppStyle.SURFACE).pack(side=tk.LEFT)

        for label, key, suffix in params:
            row = tk.Frame(card, bg=AppStyle.SURFACE)
            row.pack(fill=tk.X, pady=3)
            tk.Label(row, text=label, font=AppStyle.FONT_CAPTION, fg=AppStyle.TEXT_SECONDARY, bg=AppStyle.SURFACE, width=10, anchor="w").pack(side=tk.LEFT)

            entry_frame = tk.Frame(row, bg=AppStyle.BORDER, padx=1, pady=1)
            entry_frame.pack(side=tk.LEFT, padx=8)
            tk.Entry(entry_frame, textvariable=self.params[key], font=AppStyle.FONT_CAPTION, bg=AppStyle.SURFACE_VARIANT, fg=AppStyle.TEXT_PRIMARY, relief="flat", width=6, justify="center").pack()

            tk.Label(row, text=suffix, font=AppStyle.FONT_SMALL, fg=AppStyle.TEXT_TERTIARY, bg=AppStyle.SURFACE).pack(side=tk.LEFT)

    def _create_action_buttons(self, parent):
        frame = tk.Frame(parent, bg=AppStyle.BG_SECONDARY)
        frame.pack(fill=tk.X, pady=(8, 16))

        self.run_button = ModernButton(frame, text="実行", command=self.run_adjustment, bg_color=AppStyle.PRIMARY, width=200, height=48)
        self.run_button.pack(side=tk.LEFT, padx=(0, 12))

        self.reset_button = ModernButton(frame, text="リセット", command=self.reset_params, bg_color=AppStyle.TEXT_TERTIARY, fg_color=AppStyle.TEXT_PRIMARY, width=140, height=48)
        self.reset_button.pack(side=tk.LEFT)

        self.progress_frame = tk.Frame(parent, bg=AppStyle.BG_SECONDARY)
        self.progress_frame.pack(fill=tk.X, pady=(0, 8))
        self.progress = ttk.Progressbar(self.progress_frame, mode='indeterminate')
        self.progress.pack(fill=tk.X)

    def _create_log_area(self, parent):
        outer = tk.Frame(parent, bg=AppStyle.BORDER, padx=1, pady=1)
        outer.pack(fill=tk.BOTH, expand=True)

        card = tk.Frame(outer, bg=AppStyle.SURFACE, padx=16, pady=14)
        card.pack(fill=tk.BOTH, expand=True)

        header = tk.Frame(card, bg=AppStyle.SURFACE)
        header.pack(fill=tk.X, pady=(0, 10))
        tk.Label(header, text="ログ", font=AppStyle.FONT_BODY_BOLD, fg=AppStyle.TEXT_PRIMARY, bg=AppStyle.SURFACE).pack(side=tk.LEFT)

        log_frame = tk.Frame(card, bg=AppStyle.BG_SECONDARY, padx=1, pady=1)
        log_frame.pack(fill=tk.BOTH, expand=True)

        # スクロールバー付きのログエリア
        log_scroll = tk.Scrollbar(log_frame)
        log_scroll.pack(side=tk.RIGHT, fill=tk.Y)

        self.log_text = tk.Text(log_frame, height=12, wrap=tk.WORD, font=AppStyle.FONT_CAPTION,
                                bg=AppStyle.SURFACE_VARIANT, fg=AppStyle.TEXT_SECONDARY,
                                relief="flat", padx=12, pady=10,
                                yscrollcommand=log_scroll.set)
        self.log_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        log_scroll.config(command=self.log_text.yview)

    def _show_file_info(self, filename, file_type):
        for w in self.file_info_container.winfo_children(): w.destroy()
        # 調整パラメータセクションの前に挿入（ドロップゾーンの後）
        self.file_info_container.pack(fill=tk.X, pady=(0, 16), before=self.params_section_frame)

        styles = {"SP": (AppStyle.SP_COLOR, AppStyle.SP_BG), "SB": (AppStyle.SB_COLOR, AppStyle.SB_BG), "SD": (AppStyle.SD_COLOR, AppStyle.SD_BG)}
        color, bg = (AppStyle.PRIMARY, AppStyle.SURFACE_VARIANT)
        for k, v in styles.items():
            if k in file_type: color, bg = v; break

        outer = tk.Frame(self.file_info_container, bg=color, padx=2, pady=2)
        outer.pack(fill=tk.X)

        card = tk.Frame(outer, bg=bg, padx=20, pady=16)
        card.pack(fill=tk.X)

        tk.Label(card, text=f"読込済: {filename}", font=AppStyle.FONT_BODY_BOLD, fg=AppStyle.SUCCESS, bg=bg).pack(side=tk.LEFT)
        tk.Label(card, text=file_type, font=AppStyle.FONT_CAPTION, fg=color, bg=bg).pack(side=tk.RIGHT)

    def _update_drop_zone_success(self):
        self.drop_outer.configure(bg=AppStyle.SUCCESS)
        bg = AppStyle.SUCCESS_LIGHT
        for w in [self.drop_zone, self.drop_inner, self.drop_icon, self.drop_text, self.drop_hint]:
            w.configure(bg=bg)
        self.drop_icon.configure(text="✓")
        self.drop_text.configure(text="ファイル読込完了", fg=AppStyle.SUCCESS)
        self.drop_hint.configure(text="別のファイルをドロップして変更")

    def on_drop(self, event):
        try:
            path = event.data
            # デバッグログ
            self.log(f"ドロップイベント検出: {path}")

            # パスのクリーンアップ
            # 波括弧で囲まれている場合は削除
            if path.startswith('{') and path.endswith('}'):
                path = path[1:-1]

            # 前後の空白や引用符を削除
            path = path.strip().strip('"').strip("'")

            # パスが存在するか確認
            if os.path.exists(path):
                self.load_file(path)
            else:
                self.log(f"エラー: ファイルが見つかりません - {path}")
        except Exception as e:
            self.log(f"ドロップエラー: {str(e)}")

    def on_click_select(self, event):
        path = filedialog.askopenfilename(title="Excelファイルを選択", filetypes=[("Excel files", "*.xlsx")])
        if path: self.load_file(path)

    def load_file(self, path):
        # 読み込み中表示
        self.drop_text.configure(text="ファイル読み込み中...")
        self.root.update()

        self.full_path = path
        filename = os.path.basename(path)
        self.file_path.set(filename)

        file_format, ad_types = self.detect_file_format(path)
        self.file_format = file_format
        self.ad_types_in_file = ad_types

        # ファイルタイプ表示用の文字列を生成
        file_type_str = self.get_file_type_display_string(file_format, ad_types)
        self.file_type.set(file_type_str)

        if file_format != "unknown":
            self.file_loaded = True
            self._update_drop_zone_success()
            self._show_file_info(filename, file_type_str)
            self.log(f"\n{'='*40}")
            self.log(f"✓ ファイル読み込み完了: {filename}")
            self.log(f"形式: {file_type_str}")
            self.log(f"{'='*40}")
        else:
            # エラー時の表示を更新
            self.drop_text.configure(text="読み込みエラー", fg=AppStyle.ERROR)
            self.drop_hint.configure(text="対応ファイルを選択してください")
            self.log(f"\nエラー: 非対応ファイル - {filename}")

    def detect_file_format(self, path):
        """
        ファイル形式を検出する
        Returns:
            (format, ad_types):
                format: "unified" | "single_SP" | "single_SB" | "single_SD" | "unknown"
                ad_types: unified の場合 [has_sp, has_sb, has_sd], それ以外は None
        """
        try:
            xl = pd.ExcelFile(path)
            sheet_names = xl.sheet_names

            # デバッグ: シート名を表示
            self.log(f"検出されたシート: {', '.join(sheet_names)}")

            # 各広告タイプのシートが存在するかチェック
            # 「広告」の有無に対応（柔軟なマッチング）
            has_sp = any('スポンサープロダクト' in s and 'キャンペーン' in s for s in sheet_names)
            has_sb = any('スポンサーブランド' in s and 'キャンペーン' in s for s in sheet_names)
            has_sd = any('スポンサーディスプレイ' in s and 'キャンペーン' in s for s in sheet_names)

            self.log(f"広告タイプ検出: SP={has_sp}, SB={has_sb}, SD={has_sd}")

            count = sum([has_sp, has_sb, has_sd])

            # 複数の広告タイプが含まれる場合は統合形式
            if count >= 2:
                self.log("→ 統合形式と判定")
                return "unified", [has_sp, has_sb, has_sd]
            # 1つだけの場合は単一形式
            elif count == 1:
                if has_sp:
                    self.log("→ SP単一形式と判定")
                    return "single_SP", None
                if has_sb:
                    self.log("→ SB単一形式と判定")
                    return "single_SB", None
                if has_sd:
                    self.log("→ SD単一形式と判定")
                    return "single_SD", None

            # シート名で判定できない場合はファイル名で判定（後方互換性）
            self.log("シート名で判定できず、ファイル名で判定を試みます")
            name = os.path.basename(path).upper()
            if 'SP' in name:
                self.log("→ ファイル名からSP単一形式と判定")
                return "single_SP", None
            if 'SB' in name:
                self.log("→ ファイル名からSB単一形式と判定")
                return "single_SB", None
            if 'SD' in name:
                self.log("→ ファイル名からSD単一形式と判定")
                return "single_SD", None

            self.log("→ 判定不可（Unknown）")
            return "unknown", None
        except Exception as e:
            import traceback
            self.log(f"ファイル形式検出エラー: {str(e)}")
            self.log(f"詳細:\n{traceback.format_exc()}")
            return "unknown", None

    def get_file_type_display_string(self, file_format, ad_types):
        """表示用のファイルタイプ文字列を生成"""
        if file_format == "unified":
            has_sp, has_sb, has_sd = ad_types
            types = []
            if has_sp: types.append("SP")
            if has_sb: types.append("SB")
            if has_sd: types.append("SD")
            return f"統合形式 ({'+'.join(types)})"
        elif file_format == "single_SP":
            return "SP"
        elif file_format == "single_SB":
            return "SB"
        elif file_format == "single_SD":
            return "SD"
        else:
            return "Unknown"

    def find_sheet_name(self, path, keywords):
        """
        キーワードを含むシート名を検索する
        Args:
            path: Excelファイルのパス
            keywords: 検索するキーワードのリスト（全て含まれる必要がある）
        Returns:
            見つかったシート名、または None
        """
        try:
            xl = pd.ExcelFile(path)
            for sheet_name in xl.sheet_names:
                if all(keyword in sheet_name for keyword in keywords):
                    return sheet_name
            return None
        except:
            return None

    def detect_file_type(self, path):
        """旧バージョンとの互換性のために残す"""
        file_format, _ = self.detect_file_format(path)
        if file_format == "single_SP": return "SP"
        if file_format == "single_SB": return "SB"
        if file_format == "single_SD": return "SD"
        if file_format == "unified": return "Unified"
        return "Unknown"

    def log(self, msg):
        self.log_text.insert(tk.END, msg + "\n")
        self.log_text.see(tk.END)
        self.root.update()

    def reset_params(self):
        defaults = {
            'click_threshold_low': 10, 'click_threshold_mid': 20, 'click_threshold_high': 30,
            'reflect_rate_low': 30, 'reflect_rate_mid': 60, 'reflect_rate_high': 100,
            'reduce_rate_mid': 25, 'reduce_rate_high': 50,
            'max_change': 30, 'min_bid': 10, 'top_n': 30,
            'acos_protect_min': 30, 'acos_protect_max': 35,
            'new_kw_cpc_add_3plus': 30, 'new_kw_cpc_add_2': 15, 'new_kw_order1_max': 60,
        }
        for k, v in defaults.items(): self.params[k].set(v)
        self.log("\nパラメータをリセットしました")

    def run_adjustment(self):
        if not self.full_path:
            messagebox.showerror("エラー", "ファイルを選択してください")
            return

        self.run_button.set_state("disabled")
        self.progress.start()
        threading.Thread(target=self.execute_adjustment).start()

    def execute_adjustment(self):
        """実行メソッド（新旧形式対応版）"""
        try:
            file_format = self.file_format
            ad_types = self.ad_types_in_file

            if file_format == "unified":
                # 統合形式の処理
                self.run_unified_adjustment(ad_types)
            elif file_format.startswith("single_"):
                # 旧形式の処理（既存ロジックを維持）
                if file_format == "single_SP":
                    self.run_sp_sb_adjustment(SP_COL, SP_SEARCH_COL, "SP", sheet_mode="index", sheet_indices=(0, 1))
                elif file_format == "single_SB":
                    self.run_sp_sb_adjustment(SB_COL, SB_SEARCH_COL, "SB", sheet_mode="index", sheet_indices=(0, 1))
                elif file_format == "single_SD":
                    self.run_sd_adjustment(sheet_mode="index", sheet_index=0)
            else:
                raise ValueError("非対応のファイル形式です")
        except Exception as e:
            import traceback
            self.log(f"\nエラー: {str(e)}")
            self.log(f"詳細:\n{traceback.format_exc()}")
            messagebox.showerror("エラー", str(e))
        finally:
            self.progress.stop()
            self.run_button.set_state("normal")

    def run_unified_adjustment(self, ad_types):
        """統合形式ファイルの処理"""
        has_sp, has_sb, has_sd = ad_types
        self.log(f"\n{'='*40}")
        self.log("統合形式ファイルを検出")
        self.log(f"{'='*40}")

        # 処理対象を表示
        types_list = []
        if has_sp: types_list.append("スポンサープロダクト")
        if has_sb: types_list.append("スポンサーブランド")
        if has_sd: types_list.append("スポンサーディスプレイ")
        self.log(f"処理対象: {', '.join(types_list)}")

        output_files = []

        if has_sp:
            self.log(f"\n{'='*40}")
            self.log("【スポンサープロダクト】処理開始")
            self.log(f"{'='*40}")
            # シート名を動的に検索
            sp_campaign_sheet = self.find_sheet_name(self.full_path, ['スポンサープロダクト', 'キャンペーン'])
            sp_search_sheet = self.find_sheet_name(self.full_path, ['SP', '検索ワード'])
            self.log(f"使用するシート: {sp_campaign_sheet}, {sp_search_sheet}")
            output = self.run_sp_sb_adjustment(
                UNIFIED_SP_COL, SP_SEARCH_COL, "SP",
                sheet_mode="name",
                campaign_sheet=sp_campaign_sheet,
                search_sheet=sp_search_sheet
            )
            output_files.append(output)

        if has_sb:
            self.log(f"\n{'='*40}")
            self.log("【スポンサーブランド】処理開始")
            self.log(f"{'='*40}")
            # シート名を動的に検索
            sb_campaign_sheet = self.find_sheet_name(self.full_path, ['スポンサーブランド', 'キャンペーン'])
            sb_search_sheet = self.find_sheet_name(self.full_path, ['SB', '検索ワード'])
            self.log(f"使用するシート: {sb_campaign_sheet}, {sb_search_sheet}")
            output = self.run_sp_sb_adjustment(
                UNIFIED_SB_COL, SB_SEARCH_COL, "SB",
                sheet_mode="name",
                campaign_sheet=sb_campaign_sheet,
                search_sheet=sb_search_sheet
            )
            output_files.append(output)

        if has_sd:
            self.log(f"\n{'='*40}")
            self.log("【スポンサーディスプレイ】処理開始")
            self.log(f"{'='*40}")
            # シート名を動的に検索
            sd_campaign_sheet = self.find_sheet_name(self.full_path, ['スポンサーディスプレイ', 'キャンペーン'])
            self.log(f"使用するシート: {sd_campaign_sheet}")
            output = self.run_sd_adjustment(
                sheet_mode="name",
                campaign_sheet=sd_campaign_sheet,
                col_def=UNIFIED_SD_COL
            )
            output_files.append(output)

        self.log(f"\n{'='*40}")
        self.log("全ての処理が完了しました")
        self.log(f"{'='*40}")
        for f in output_files:
            self.log(f"出力: {os.path.basename(f)}")

        file_list = "\n".join([os.path.basename(f) for f in output_files])
        messagebox.showinfo("完了", f"処理が完了しました\n\n出力ファイル:\n{file_list}")

    def safe_float(self, value, default=0.0):
        if pd.isna(value) or value == '' or value == '-': return default
        try:
            if isinstance(value, str):
                value = value.replace('%', '').replace(',', '').replace('yen', '').strip()
            return float(value)
        except: return default

    def safe_int(self, value, default=""):
        """整数値を安全に取得（数値変換できない場合はdefaultを返す）"""
        if pd.isna(value) or value == '' or value == '-':
            return default
        try:
            if isinstance(value, str):
                # '上限なし'などの文字列はdefaultを返す
                value = value.replace(',', '').strip()
                if not value.replace('.', '').replace('-', '').isdigit():
                    return default
            return str(int(float(value)))
        except:
            return default

    def apply_limit(self, current, new_bid, max_change):
        diff = new_bid - current
        if diff > max_change: return current + max_change, True
        if diff < -max_change: return current - max_change, True
        return new_bid, False

    def run_sp_sb_adjustment(self, col, search_col, ad_type, sheet_mode="index", sheet_indices=None, campaign_sheet=None, search_sheet=None):
        """
        SP/SBの調整処理（新旧形式対応版）

        Args:
            sheet_mode: "index" (旧形式、インデックス指定) or "name" (新形式、シート名指定)
            sheet_indices: (campaign_index, search_index) - sheet_mode="index"の場合に使用
            campaign_sheet: キャンペーンシート名 - sheet_mode="name"の場合に使用
            search_sheet: 検索ワードシート名 - sheet_mode="name"の場合に使用
        """
        self.log(f"\n{'='*40}")
        self.log(f"{ad_type}の調整を開始")

        # シートの読み込み
        if sheet_mode == "index":
            campaign_idx, search_idx = sheet_indices
            df_camp = pd.read_excel(self.full_path, sheet_name=campaign_idx)
            df_search = pd.read_excel(self.full_path, sheet_name=search_idx)
        else:  # sheet_mode == "name"
            df_camp = pd.read_excel(self.full_path, sheet_name=campaign_sheet)
            df_search = pd.read_excel(self.full_path, sheet_name=search_sheet)

        # B列（エンティティ）でフィルタリング - SP/SBの場合
        original_rows = len(df_camp)
        if ad_type in ['SP', 'SB']:
            allowed_entities = ['キャンペーン', 'キーワード', '商品ターゲティング']
            # B列は0-indexedで列1
            df_camp = df_camp[df_camp.iloc[:, 1].isin(allowed_entities)].reset_index(drop=True)
            filtered_rows = len(df_camp)
            self.log(f"B列エンティティフィルタリング: {original_rows}行 → {filtered_rows}行（{allowed_entities}のみ）")

        self.log(f"キャンペーン: {len(df_camp)}行")
        self.log(f"検索ワード: {len(df_search)}行")

        p = {k: v.get() for k, v in self.params.items()}

        search_cv = {}
        target_acos_map = {}
        original_bid_map = {}
        search_term_data = []
        existing_keywords = set()  # 既存キーワード/ターゲティングを収集

        # まず、キャンペーンシートからtarget_acos_mapとoriginal_bid_mapを構築
        # （検索ワードシートには許容ACOSが存在しない）
        # 同時に既存キーワード/ターゲティングを収集
        for _, row in df_camp.iterrows():
            kw_id = self.safe_int(row.iloc[col['KEYWORD_ID']])
            tg_id = self.safe_int(row.iloc[col['TARGETING_ID']])
            id_key = kw_id if kw_id and kw_id != 'nan' else tg_id

            if not id_key or id_key == 'nan':
                continue

            # TARGET_ACOS列の読み取り
            target_acos_val = self.safe_float(row.iloc[col['TARGET_ACOS']], 0)
            if target_acos_val > 0 and id_key not in target_acos_map:
                target_acos_map[id_key] = target_acos_val

            # 元の入札額を保存
            bid_val = self.safe_float(row.iloc[col['BID']], 0)
            if bid_val > 0 and id_key not in original_bid_map:
                original_bid_map[id_key] = bid_val

            # 既存キーワード/ターゲティングを収集（ポートフォリオ単位）
            portfolio = str(row.iloc[col['PORTFOLIO']]).strip() if pd.notna(row.iloc[col['PORTFOLIO']]) else ""
            # 'nan'文字列を空文字列に変換
            if portfolio.lower() == 'nan':
                portfolio = ""

            # キーワードテキスト
            if 'KEYWORD_TEXT' in col:
                kw_text = str(row.iloc[col['KEYWORD_TEXT']]).strip().lower() if pd.notna(row.iloc[col['KEYWORD_TEXT']]) else ""
                if kw_text and kw_text != 'nan':
                    existing_keywords.add((kw_text, portfolio))

            # 商品ターゲティング式
            if 'PRODUCT_TARGETING' in col:
                pt_expr = str(row.iloc[col['PRODUCT_TARGETING']]).strip().lower() if pd.notna(row.iloc[col['PRODUCT_TARGETING']]) else ""
                if pt_expr and pt_expr != 'nan':
                    existing_keywords.add((pt_expr, portfolio))

        self.log(f"既存キーワード/ターゲティング: {len(existing_keywords)}件")

        # マニュアルキャンペーン情報を収集（ポートフォリオごと）
        manual_campaigns_by_portfolio = {}
        if 'TARGETING_TYPE' in col and 'MATCH_TYPE' in col and 'CAMPAIGN_ID' in col:
            for _, row in df_camp.iterrows():
                portfolio = str(row.iloc[col['PORTFOLIO']]).strip() if pd.notna(row.iloc[col['PORTFOLIO']]) else ""
                if portfolio.lower() == 'nan':
                    portfolio = ""

                targeting_type = str(row.iloc[col['TARGETING_TYPE']]).strip() if pd.notna(row.iloc[col['TARGETING_TYPE']]) else ""

                # マニュアルターゲティングのみ処理
                if targeting_type == 'マニュアル':
                    # ポートフォリオ単位で初期化
                    if portfolio not in manual_campaigns_by_portfolio:
                        manual_campaigns_by_portfolio[portfolio] = {
                            'campaign_ids': set(),
                            'exact_match_keywords': set()
                        }

                    # キャンペーンIDを収集
                    campaign_id = self.safe_int(row.iloc[col['CAMPAIGN_ID']])
                    if campaign_id and campaign_id != 'nan':
                        manual_campaigns_by_portfolio[portfolio]['campaign_ids'].add(campaign_id)

                    # 完全一致キーワードを収集
                    match_type = str(row.iloc[col['MATCH_TYPE']]).strip() if pd.notna(row.iloc[col['MATCH_TYPE']]) else ""
                    if match_type == '完全一致' and 'KEYWORD_TEXT' in col:
                        kw_text = str(row.iloc[col['KEYWORD_TEXT']]).strip().lower() if pd.notna(row.iloc[col['KEYWORD_TEXT']]) else ""
                        if kw_text and kw_text != 'nan':
                            manual_campaigns_by_portfolio[portfolio]['exact_match_keywords'].add(kw_text)

            self.log(f"マニュアルキャンペーン情報: {len(manual_campaigns_by_portfolio)}ポートフォリオ")
            for pf, info in manual_campaigns_by_portfolio.items():
                self.log(f"  {pf or '(ポートフォリオなし)'}: キャンペーン{len(info['campaign_ids'])}件, 完全一致キーワード{len(info['exact_match_keywords'])}件")

        # 検索ワードシートからCV情報を取得
        for _, row in df_search.iterrows():
            kw_id = self.safe_int(row.iloc[search_col['KEYWORD_ID']])
            tg_id = self.safe_int(row.iloc[search_col['TARGETING_ID']])
            id_key = kw_id if kw_id and kw_id != 'nan' else tg_id

            if not id_key or id_key == 'nan': continue

            orders = self.safe_float(row.iloc[search_col['ORDERS']], 0)
            clicks = self.safe_float(row.iloc[search_col['CLICKS']], 0)

            if id_key not in search_cv:
                search_cv[id_key] = {'has_cv': False, 'total_clicks': 0, 'cv_terms': 0}

            search_cv[id_key]['total_clicks'] += clicks
            if orders > 0:
                search_cv[id_key]['has_cv'] = True
                search_cv[id_key]['cv_terms'] += 1

            portfolio = str(row.iloc[search_col['PORTFOLIO']]).strip() if pd.notna(row.iloc[search_col['PORTFOLIO']]) else ""
            # 'nan'文字列を空文字列に変換
            if portfolio.lower() == 'nan':
                portfolio = ""
            search_term = str(row.iloc[search_col['SEARCH_TERM']]) if pd.notna(row.iloc[search_col['SEARCH_TERM']]) else ""
            spend = self.safe_float(row.iloc[search_col['SPEND']], 0)
            sales = self.safe_float(row.iloc[search_col['SALES']], 0)

            # キャンペーンID・グループID・名前を取得
            campaign_id = self.safe_int(row.iloc[search_col['CAMPAIGN_ID']])
            ad_group_id = self.safe_int(row.iloc[search_col['AD_GROUP_ID']])
            campaign_name = str(row.iloc[search_col['CAMPAIGN_NAME']]) if pd.notna(row.iloc[search_col['CAMPAIGN_NAME']]) else ""
            ad_group_name = str(row.iloc[search_col['AD_GROUP_NAME']]) if pd.notna(row.iloc[search_col['AD_GROUP_NAME']]) else ""

            if search_term and search_term != 'nan':
                search_term_data.append({
                    'search_term': search_term, 'portfolio': portfolio, 'id_key': id_key,
                    'campaign_id': campaign_id, 'ad_group_id': ad_group_id,
                    'campaign_name': campaign_name, 'ad_group_name': ad_group_name,
                    'clicks': clicks, 'spend': spend, 'sales': sales, 'orders': orders
                })

        df_sorted = df_camp.copy()
        df_sorted['_sales'] = df_sorted.iloc[:, col['SALES']].apply(lambda x: self.safe_float(x, 0))
        top_n_idx = set(df_sorted.sort_values('_sales', ascending=False).head(p['top_n']).index)

        new_bids, original_bids, reasons = [], [], []
        stats = {}

        for idx, row in df_camp.iterrows():
            kw_id = self.safe_int(row.iloc[col['KEYWORD_ID']])
            tg_id = self.safe_int(row.iloc[col['TARGETING_ID']])
            id_key = kw_id if kw_id and kw_id != 'nan' else tg_id

            current = self.safe_float(row.iloc[col['BID']], 0)
            clicks = self.safe_float(row.iloc[col['CLICKS']], 0)
            acos = self.safe_float(row.iloc[col['ACOS']], 0)
            target_acos = self.safe_float(row.iloc[col['TARGET_ACOS']], 0)

            original_bids.append(current)

            if target_acos == 0 and id_key in target_acos_map:
                target_acos = target_acos_map[id_key]

            cv_info = search_cv.get(id_key, {'has_cv': False, 'cv_terms': 0})
            has_cv = cv_info['has_cv']
            cv_terms = cv_info['cv_terms']

            reason, new_bid = self._calculate_bid(
                current, clicks, acos, target_acos, has_cv, cv_terms, idx in top_n_idx, p
            )

            new_bids.append(new_bid)
            reasons.append(reason)
            # 理由をそのままキーとして集計
            stats[reason] = stats.get(reason, 0) + 1

        # カテゴリ別に整理してログ出力
        self._log_organized_summary(stats)

        new_kw_candidates, skipped_existing = self._find_new_keywords(
            search_term_data, target_acos_map, original_bid_map, p, existing_keywords, manual_campaigns_by_portfolio
        )
        exclude_candidates = self._find_exclude_keywords(search_term_data, p)

        self.log(f"\n【キーワード候補】")
        self.log(f"  新規キーワード候補: {len(new_kw_candidates)}件")
        self.log(f"  既存キーワード除外: {skipped_existing}件")
        self.log(f"  除外キーワード候補: {len(exclude_candidates)}件")

        # 出力ファイル名の生成
        today = datetime.now().strftime('%Y%m%d')
        if sheet_mode == "index":
            # 旧形式: ファイル名に _adjusted を追加
            output = self.full_path.replace('.xlsx', '_adjusted.xlsx')
        else:
            # 新形式: スポンサープロダクト_YYYYMMDD.xlsx 形式
            ad_type_jp = {
                'SP': 'スポンサープロダクト',
                'SB': 'スポンサーブランド'
            }
            dir_path = os.path.dirname(self.full_path)
            filename = f"{ad_type_jp.get(ad_type, ad_type)}_{today}.xlsx"
            output = os.path.join(dir_path, filename)

        # 新しいワークブックを作成
        wb = Workbook()
        ws = wb.active
        ws.title = campaign_sheet if sheet_mode == "name" else f"{ad_type}キャンペーン"

        # フィルタリング済みのdf_campからデータを書き込み
        # 1. ヘッダー行を書き込み
        for c_idx, col_name in enumerate(df_camp.columns, 1):
            ws.cell(row=1, column=c_idx, value=col_name)

        # 2. データ行を書き込み（フィルタリング済み、数式は値として書き込まれる）
        for r_idx, (_, row) in enumerate(df_camp.iterrows(), 2):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        # 3. 調整結果列を追加（df_campの列数+1列目から開始）
        out_col = len(df_camp.columns) + 1
        self.log(f"デバッグ: データフレーム列数={len(df_camp.columns)}, 出力開始列={out_col}")
        ws.cell(row=1, column=out_col, value="元の入札額")
        ws.cell(row=1, column=out_col+1, value="新入札額")
        ws.cell(row=1, column=out_col+2, value="理由")

        for i, (o, n, r) in enumerate(zip(original_bids, new_bids, reasons), 2):
            ws.cell(row=i, column=out_col, value=o)
            ws.cell(row=i, column=out_col+1, value=n)
            ws.cell(row=i, column=out_col+2, value=r)

        # 新規キーワード候補シートを追加
        if new_kw_candidates:
            ws_new = wb.create_sheet("新規キーワード候補")
            ws_new.append(list(new_kw_candidates[0].keys()))
            for c in new_kw_candidates:
                ws_new.append(list(c.values()))

        # 除外候補シートを追加
        if exclude_candidates:
            ws_ex = wb.create_sheet("除外候補")
            ws_ex.append(list(exclude_candidates[0].keys()))
            for c in exclude_candidates:
                ws_ex.append(list(c.values()))

        wb.save(output)

        self.log(f"\n完了: {os.path.basename(output)}")

        return output

    def _calculate_bid(self, current, clicks, acos, target_acos, has_cv, cv_terms, is_top, p):
        min_bid = p['min_bid']
        max_change = p['max_change']

        # 1. 許容ACOS未設定
        if target_acos == 0:
            return "許容ACOS未設定", current

        # 2. クリック数10以下
        if clicks <= p['click_threshold_low']:
            return f"クリック{p['click_threshold_low']}以下（データ不足）", current

        # 3. ACOS=0（売上なし）+ クリック11〜20 → 10%削減（制限なし）
        if acos == 0 and p['click_threshold_low'] < clicks <= p['click_threshold_mid']:
            new = max(min_bid, round(current * 0.90))
            return f"売上なし+クリック{p['click_threshold_low']+1}-{p['click_threshold_mid']}→10%削減", new

        # 4. ACOS=0（売上なし）+ クリック21〜29 → 25%削減（制限なし）
        if acos == 0 and p['click_threshold_mid'] < clicks < p['click_threshold_high']:
            new = max(min_bid, round(current * 0.75))
            return f"売上なし+クリック{p['click_threshold_mid']+1}-{p['click_threshold_high']-1}→25%削減", new

        # 5. ACOS=0（売上なし）+ クリック30以上 → 50%削減（制限なし）
        if acos == 0 and clicks >= p['click_threshold_high']:
            new = max(min_bid, round(current * 0.50))
            return f"売上なし+クリック{p['click_threshold_high']}以上→50%削減", new

        # 6. 売上上位N位 + ACOS 30%〜35% → 現状維持
        acos_min = p['acos_protect_min'] / 100
        acos_max = p['acos_protect_max'] / 100
        if is_top and acos_min <= acos <= acos_max:
            return f"上位商品・ACOS適正({acos*100:.1f}%)、現状維持", current

        if acos == 0:
            return "ACOS計算不可", current

        # 基本計算
        ratio = target_acos / acos
        calculated = current * ratio
        adjustment = calculated - current

        # 7. クリック11〜20件（売上あり）→ 30%反映
        if p['click_threshold_low'] < clicks <= p['click_threshold_mid']:
            new = current + (adjustment * p['reflect_rate_low'] / 100)
            new, limited = self.apply_limit(current, new, max_change)
            new = max(min_bid, round(new))
            direction = "↑" if new > current else "↓" if new < current else "→"
            limit_note = "（±30円制限）" if limited else ""
            return f"クリック{p['click_threshold_low']+1}-{p['click_threshold_mid']}（{p['reflect_rate_low']}%反映）、入札{direction}{limit_note}", new

        # 8. クリック21〜29件（売上あり）→ 60%反映
        if p['click_threshold_mid'] < clicks < p['click_threshold_high']:
            new = current + (adjustment * p['reflect_rate_mid'] / 100)
            new, limited = self.apply_limit(current, new, max_change)
            new = max(min_bid, round(new))
            direction = "↑" if new > current else "↓" if new < current else "→"
            limit_note = "（±30円制限）" if limited else ""
            return f"クリック{p['click_threshold_mid']+1}-{p['click_threshold_high']-1}（{p['reflect_rate_mid']}%反映）、入札{direction}{limit_note}", new

        # 9. 検索語CV無し + クリック30以上 → 50%削減
        if not has_cv and clicks >= p['click_threshold_high']:
            new = current * 0.50
            new, limited = self.apply_limit(current, new, max_change)
            new = max(min_bid, round(new))
            limit_note = "（±30円制限）" if limited else ""
            return f"検索語CV無し+クリック{p['click_threshold_high']}以上→50%削減{limit_note}", new

        # 10. 通常調整（クリック30以上・CV有り）
        if clicks >= p['click_threshold_high']:
            new, limited = self.apply_limit(current, calculated, max_change)
            new = max(min_bid, round(new))
            direction = "↑" if new > current else "↓" if new < current else "→"
            limit_note = "（±30円制限）" if limited else ""
            return f"通常調整（クリック{p['click_threshold_high']}以上）、入札{direction}{limit_note}", new

        # その他
        return "その他", current

    def _find_new_keywords(self, data, target_acos_map, original_bid_map, p, existing_keywords, manual_campaigns_by_portfolio=None):
        portfolio_terms = {}
        for item in data:
            key = (item['search_term'], item['portfolio'])
            if key not in portfolio_terms:
                portfolio_terms[key] = []
            portfolio_terms[key].append(item)

        candidates = []
        skipped_existing = 0  # 既存キーワードとしてスキップした件数
        skipped_manual = 0  # マニュアルキャンペーンに存在するためスキップした件数

        for (term, portfolio), items in portfolio_terms.items():
            # 検索語句を正規化（大文字小文字を区別しない）
            term_normalized = term.strip().lower()
            # ポートフォリオも正規化
            portfolio_normalized = portfolio.strip() if portfolio else ""
            if portfolio_normalized.lower() == 'nan':
                portfolio_normalized = ""

            # 既存キーワードチェック（従来のロジック）
            if (term_normalized, portfolio_normalized) in existing_keywords:
                skipped_existing += 1
                continue  # 既存のため除外

            # マニュアルキャンペーンの完全一致キーワードチェック（新ロジック）
            if manual_campaigns_by_portfolio and portfolio_normalized in manual_campaigns_by_portfolio:
                manual_info = manual_campaigns_by_portfolio[portfolio_normalized]
                # このポートフォリオのマニュアルキャンペーンに完全一致で存在するかチェック
                if term_normalized in manual_info['exact_match_keywords']:
                    skipped_manual += 1
                    continue  # マニュアルキャンペーンに既に存在するため除外

            if len(items) == 1 and items[0]['orders'] > 0:
                item = items[0]
                cpc = item['spend'] / item['clicks'] if item['clicks'] > 0 else 0
                # ACOSは小数で計算（例: 0.282 = 28.2%）
                acos = item['spend'] / item['sales'] if item['sales'] > 0 else 0
                cvr = item['orders'] / item['clicks'] if item['clicks'] > 0 else 0
                # target_acos_mapの値はキャンペーンシートから取得（小数形式: 0.30 = 30%）
                target = target_acos_map.get(item['id_key'], 0)
                # 元の入札額を取得
                original_bid = original_bid_map.get(item['id_key'], 0)

                rec_bid = "N/A"
                target_str = "N/A"
                action = "N/A"
                orders = int(item['orders'])
                # 両方小数で計算（target/acos比率）
                if target > 0 and acos > 0:
                    raw_rec = round(cpc * (target / acos))
                    rec = max(raw_rec, p['min_bid'])
                    # 入札上限ルール: 注文数に応じてCPC上乗せ上限を変える
                    cpc_rounded = round(cpc)
                    cpc_add_3plus = p.get('new_kw_cpc_add_3plus', 30)
                    cpc_add_2 = p.get('new_kw_cpc_add_2', 15)
                    order1_max = p.get('new_kw_order1_max', 60)

                    # 方向判定（CPC基準）
                    if rec > cpc_rounded:
                        direction = "↑"
                    elif rec < cpc_rounded:
                        direction = "↓"
                    else:
                        direction = "→"

                    if orders >= 3:
                        # 注文3件以上: CPC + 30円まで
                        if rec > cpc_rounded + cpc_add_3plus:
                            rec = cpc_rounded + cpc_add_3plus
                            action = f"注文{orders}件・{direction}上限+{cpc_add_3plus}円"
                        else:
                            diff = rec - cpc_rounded
                            if diff > 0:
                                action = f"注文{orders}件・{direction}+{diff}円"
                            elif diff < 0:
                                action = f"注文{orders}件・{direction}{diff}円"
                            else:
                                action = f"注文{orders}件・維持"
                    elif orders == 2:
                        # 注文2件: CPC + 15円まで
                        if rec > cpc_rounded + cpc_add_2:
                            rec = cpc_rounded + cpc_add_2
                            action = f"注文2件・{direction}上限+{cpc_add_2}円"
                        else:
                            diff = rec - cpc_rounded
                            if diff > 0:
                                action = f"注文2件・{direction}+{diff}円"
                            elif diff < 0:
                                action = f"注文2件・{direction}{diff}円"
                            else:
                                action = f"注文2件・維持"
                    else:
                        # 注文1件: CPC + 15円まで、かつ最大60円
                        hit_cpc_limit = False
                        hit_max_limit = False
                        if rec > cpc_rounded + cpc_add_2:
                            rec = cpc_rounded + cpc_add_2
                            hit_cpc_limit = True
                        if rec > order1_max:
                            rec = order1_max
                            hit_max_limit = True

                        diff = rec - cpc_rounded
                        if hit_max_limit:
                            action = f"注文1件・{direction}上限{order1_max}円"
                        elif hit_cpc_limit:
                            action = f"注文1件・{direction}上限+{cpc_add_2}円"
                        elif diff > 0:
                            action = f"注文1件・{direction}+{diff}円"
                        elif diff < 0:
                            action = f"注文1件・{direction}{diff}円"
                        else:
                            action = f"注文1件・維持"

                    rec_bid = f"{rec}円"
                    # キャンペーンシートの許容ACOSは小数形式（0.30 = 30%）
                    target_str = f"{target*100:.0f}%"

                # ACOSは計算値なので常に小数
                acos_str = f"{acos*100:.1f}%"

                # マニュアルキャンペーンIDを取得
                manual_campaign_id = 'マニュアルターゲティングなし'
                if manual_campaigns_by_portfolio and portfolio_normalized in manual_campaigns_by_portfolio:
                    manual_info = manual_campaigns_by_portfolio[portfolio_normalized]
                    if manual_info['campaign_ids']:
                        # 複数ある場合は最初のIDを使用（通常は1つのはず）
                        manual_campaign_id = ', '.join(sorted(manual_info['campaign_ids']))

                candidates.append({
                    'カスタマー検索語': term, 'ポートフォリオ名': portfolio,
                    'オートキャンペーンID': item.get('campaign_id', ''),
                    'マニュアルキャンペーンID': manual_campaign_id,
                    '広告グループID': item.get('ad_group_id', ''),
                    'キャンペーン名': item.get('campaign_name', ''),
                    '広告グループ名': item.get('ad_group_name', ''),
                    '元ターゲティングID': item['id_key'], 'クリック数': int(item['clicks']),
                    '売上': round(item['sales']), '注文数': int(item['orders']),
                    'CPC': f"{cpc:.1f}円", 'ACOS': acos_str,
                    'コンバージョン率': f"{cvr*100:.2f}%",
                    '推奨入札単価': rec_bid, '許容ACOS': target_str,
                    '推奨アクション': action
                })

        sorted_candidates = sorted(candidates, key=lambda x: x['クリック数'], reverse=True)

        # ログ出力（デバッグ用）
        if manual_campaigns_by_portfolio:
            self.log(f"  既存キーワード除外: {skipped_existing}件")
            self.log(f"  マニュアル完全一致除外: {skipped_manual}件")

        return sorted_candidates, skipped_existing + skipped_manual

    def _find_exclude_keywords(self, data, p):
        portfolio_terms = {}
        for item in data:
            key = (item['search_term'], item['portfolio'])
            if key not in portfolio_terms:
                portfolio_terms[key] = []
            portfolio_terms[key].append(item)

        candidates = []
        for (term, portfolio), items in portfolio_terms.items():
            total_clicks = sum(i['clicks'] for i in items)
            total_orders = sum(i['orders'] for i in items)

            if total_clicks >= p['click_threshold_high'] and total_orders == 0:
                total_spend = sum(i['spend'] for i in items)
                cpc = total_spend / total_clicks if total_clicks > 0 else 0
                related_ids = list(set([i['id_key'] for i in items if i['id_key'] and i['id_key'] != 'nan']))

                # 最初のアイテムからキャンペーン・グループ情報を取得
                first_item = items[0]
                campaign_id = first_item.get('campaign_id', '')
                ad_group_id = first_item.get('ad_group_id', '')
                campaign_name = first_item.get('campaign_name', '')
                ad_group_name = first_item.get('ad_group_name', '')

                candidates.append({
                    'カスタマー検索語': term, 'ポートフォリオ名': portfolio,
                    'キャンペーンID': campaign_id,
                    '広告グループID': ad_group_id,
                    'キャンペーン名': campaign_name,
                    '広告グループ名': ad_group_name,
                    '関連ターゲティング数': len(related_ids),
                    '関連ターゲティングID': ', '.join(related_ids[:5]),
                    '合計クリック数': int(total_clicks), '合計支出': f"{total_spend:.0f}円",
                    'CPC': f"{cpc:.1f}円", 'ACOS': 'N/A', 'コンバージョン率': "0.00%",
                    '推奨アクション': '除外キーワード追加検討'
                })

        return sorted(candidates, key=lambda x: x['合計クリック数'], reverse=True)

    def _log_organized_summary(self, stats):
        """調整結果をカテゴリ別に整理してログ出力"""
        self.log("\n" + "="*50)
        self.log("調整結果サマリー")
        self.log("="*50)

        # カテゴリ定義（表示順序）
        categories = {
            'データ不足': [],
            '売上なし（ACOS=0）': [],
            '売上あり・通常調整': [],
            '売上あり・部分反映': [],
            '検索語CV関連': [],
            '上位商品保護': [],
            '設定・その他': [],
        }

        # 理由をカテゴリに分類
        for reason, count in stats.items():
            if 'データ不足' in reason or 'クリック10以下' in reason:
                categories['データ不足'].append((reason, count))
            elif '売上なし' in reason or 'ACOS=0' in reason:
                categories['売上なし（ACOS=0）'].append((reason, count))
            elif '通常調整' in reason:
                categories['売上あり・通常調整'].append((reason, count))
            elif '反映' in reason:
                categories['売上あり・部分反映'].append((reason, count))
            elif '検索語CV' in reason:
                categories['検索語CV関連'].append((reason, count))
            elif '上位商品' in reason or '現状維持' in reason:
                categories['上位商品保護'].append((reason, count))
            else:
                categories['設定・その他'].append((reason, count))

        # 合計を計算
        total = sum(stats.values())

        # カテゴリ別に出力
        for cat_name, items in categories.items():
            if items:
                cat_total = sum(c for _, c in items)
                self.log(f"\n【{cat_name}】 計{cat_total}件")

                # 入札↑、入札→、入札↓の順でソート
                def sort_key(item):
                    reason = item[0]
                    if '↑' in reason:
                        return (0, -item[1])
                    elif '→' in reason:
                        return (1, -item[1])
                    elif '↓' in reason:
                        return (2, -item[1])
                    else:
                        return (3, -item[1])

                for reason, count in sorted(items, key=sort_key):
                    self.log(f"  {reason}: {count}件")

        self.log(f"\n{'─'*50}")
        self.log(f"合計: {total}件")

    def run_sd_adjustment(self, sheet_mode="index", sheet_index=None, campaign_sheet=None, col_def=None):
        """
        SDの調整処理（新旧形式対応版）

        Args:
            sheet_mode: "index" (旧形式、インデックス指定) or "name" (新形式、シート名指定)
            sheet_index: キャンペーンシートのインデックス - sheet_mode="index"の場合に使用
            campaign_sheet: キャンペーンシート名 - sheet_mode="name"の場合に使用
            col_def: 列定義辞書（統合形式の場合はUNIFIED_SD_COL、旧形式の場合はNoneでSD_COL使用）
        """
        self.log(f"\n{'='*40}")
        self.log("SDの調整を開始")

        # シートの読み込み
        if sheet_mode == "index":
            df = pd.read_excel(self.full_path, sheet_name=sheet_index)
        else:  # sheet_mode == "name"
            df = pd.read_excel(self.full_path, sheet_name=campaign_sheet)

        # B列（エンティティ）でフィルタリング - SDの場合
        original_rows = len(df)
        allowed_entities = ['キャンペーン', 'オーディエンスターゲティング', 'コンテキストターゲティング']
        # B列は0-indexedで列1
        df = df[df.iloc[:, 1].isin(allowed_entities)].reset_index(drop=True)
        filtered_rows = len(df)
        self.log(f"B列エンティティフィルタリング: {original_rows}行 → {filtered_rows}行（{allowed_entities}のみ）")

        self.log(f"キャンペーン: {len(df)}行")

        p = {k: v.get() for k, v in self.params.items()}
        col = col_def if col_def is not None else SD_COL

        df_sorted = df.copy()
        df_sorted['_sales'] = df_sorted.iloc[:, col['SALES']].apply(lambda x: self.safe_float(x, 0))
        top_n_idx = set(df_sorted.sort_values('_sales', ascending=False).head(p['top_n']).index)

        new_bids, original_bids, reasons = [], [], []
        stats = {}

        for idx, row in df.iterrows():
            current = self.safe_float(row.iloc[col['BID']], 0)
            clicks = self.safe_float(row.iloc[col['CLICKS']], 0)
            acos = self.safe_float(row.iloc[col['ACOS']], 0)
            target_acos = self.safe_float(row.iloc[col['TARGET_ACOS']], 0)

            original_bids.append(current)

            reason, new_bid = self._calculate_bid(
                current, clicks, acos, target_acos, True, 3, idx in top_n_idx, p
            )

            new_bids.append(new_bid)
            reasons.append(reason)
            # 理由をそのままキーとして集計
            stats[reason] = stats.get(reason, 0) + 1

        # カテゴリ別に整理してログ出力
        self._log_organized_summary(stats)

        # 出力ファイル名の生成
        today = datetime.now().strftime('%Y%m%d')
        if sheet_mode == "index":
            # 旧形式: ファイル名に _adjusted を追加
            output = self.full_path.replace('.xlsx', '_adjusted.xlsx')
        else:
            # 新形式: スポンサーディスプレイ_YYYYMMDD.xlsx 形式
            dir_path = os.path.dirname(self.full_path)
            filename = f"スポンサーディスプレイ_{today}.xlsx"
            output = os.path.join(dir_path, filename)

        # 新しいワークブックを作成
        wb = Workbook()
        ws = wb.active
        ws.title = campaign_sheet if sheet_mode == "name" else "SDキャンペーン"

        # フィルタリング済みのdfからデータを書き込み
        # 1. ヘッダー行を書き込み
        for c_idx, col_name in enumerate(df.columns, 1):
            ws.cell(row=1, column=c_idx, value=col_name)

        # 2. データ行を書き込み（フィルタリング済み、数式は値として書き込まれる）
        for r_idx, (_, row) in enumerate(df.iterrows(), 2):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        # 3. 調整結果列を追加（dfの列数+1列目から開始）
        out_col = len(df.columns) + 1
        self.log(f"デバッグ: データフレーム列数={len(df.columns)}, 出力開始列={out_col}")
        ws.cell(row=1, column=out_col, value="元の入札額")
        ws.cell(row=1, column=out_col+1, value="新入札額")
        ws.cell(row=1, column=out_col+2, value="理由")

        for i, (o, n, r) in enumerate(zip(original_bids, new_bids, reasons), 2):
            ws.cell(row=i, column=out_col, value=o)
            ws.cell(row=i, column=out_col+1, value=n)
            ws.cell(row=i, column=out_col+2, value=r)

        wb.save(output)

        self.log(f"\n完了: {os.path.basename(output)}")

        return output


if __name__ == "__main__":
    root = TkinterDnD.Tk()
    app = AmazonBidAdjusterApp(root)
    root.mainloop()

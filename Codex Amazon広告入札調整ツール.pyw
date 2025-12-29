# -*- coding: utf-8 -*-
"""
Amazon広告 入札単価自動調整ツール v13
GUI版 - シンプル＆クリーン版
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from tkinterdnd2 import DND_FILES, TkinterDnD
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os
import re
import threading


# ==================== 列インデックス定義（固定）====================
# Amazonレポートの列は固定形式なので、動的検出は不要

# SPキャンペーンシート
SP_COL = {
    'ENTITY': 1,          # B列
    'CAMPAIGN_ID': 3,    # キャンペーンID
    'AD_GROUP_ID': 4,    # 広告グループID
    'KEYWORD_ID': 7,      # H列
    'TARGETING_ID': 8,    # I列
    'PORTFOLIO': 13,      # N列
    'BID': 27,            # AB列
    'KEYWORD_TEXT': 28,   # AC列（キーワードテキスト）
    'MATCH_TYPE': 31,     # AF列（マッチタイプ）
    'PRODUCT_TARGETING': 35,  # AJ列（商品ターゲティング式）
    'CLICKS': 42,         # クリック数
    'SPEND': 44,          # 支出
    'SALES': 45,          # 売上
    'ORDERS': 46,         # 注文
    'ACOS': 49,           # ACOS
    'TARGET_ACOS': 53,    # BB列
    'CAMPAIGN_NAME_INFO': 11, # L列（キャンペーン名（情報提供のみ））
}

# SP検索ワードシート（許容ACOSは検索ワードシートには存在しない）
SP_SEARCH_COL = {
    'CAMPAIGN_ID': 1,     # B列（キャンペーンID）
    'AD_GROUP_ID': 2,     # C列（広告グループID）
    'KEYWORD_ID': 3,      # D列
    'TARGETING_ID': 4,    # E列
    'CAMPAIGN_NAME': 5,   # F列（キャンペーン名）
    'AD_GROUP_NAME': 6,   # G列（広告グループ名）
    'PORTFOLIO': 7,       # H列
    'KEYWORD_TEXT': 11,   # L列（キーワードテキスト）
    'MATCH_TYPE': 12,     # M列（マッチタイプ）
    'SEARCH_TERM': 15,    # P列
    'CLICKS': 17,         # R列
    'SPEND': 19,          # T列
    'SALES': 20,          # U列
    'ORDERS': 21,         # V列
    # TARGET_ACOSは検索ワードシートには存在しない - キャンペーンシートから取得する
}

# SBキャンペーンシート
SB_COL = {
    'CAMPAIGN_ID': 3,    # キャンペーンID
    'KEYWORD_ID': 7,      # H列
    'TARGETING_ID': 8,    # I列
    'PORTFOLIO': 11,      # L列（ポートフォリオ名）
    'BID': 21,            # V列（入札額）
    'KEYWORD_TEXT': 22,   # W列（キーワードテキスト）
    'PRODUCT_TARGETING': 24,  # Y列（商品ターゲティング式）
    'CLICKS': 41,         # AP列
    'SPEND': 43,          # AR列（支出）
    'SALES': 44,          # AS列
    'ORDERS': 45,         # AT列
    'ACOS': 48,           # AW列
    'TARGET_ACOS': 51,    # AZ列（許容ACOS）
}

# SB検索ワードシート（許容ACOSは検索ワードシートには存在しない）
SB_SEARCH_COL = {
    'CAMPAIGN_ID': 1,     # B列（キャンペーンID）
    'AD_GROUP_ID': 2,     # C列（広告グループID）
    'KEYWORD_ID': 3,      # D列
    'TARGETING_ID': 4,    # E列
    'CAMPAIGN_NAME': 5,   # F列（キャンペーン名）
    'AD_GROUP_NAME': 6,   # G列（広告グループ名）
    'PORTFOLIO': 5,       # F列（SB検索ワードシートにはポートフォリオ列がないためキャンペーン名を使用）
    'SEARCH_TERM': 13,    # N列（カスタマー検索用語）
    'CLICKS': 15,         # P列
    'SPEND': 17,          # R列（支出）
    'SALES': 18,          # S列
    'ORDERS': 19,         # T列
    # TARGET_ACOSは検索ワードシートには存在しない - キャンペーンシートから取得する
}

# SDキャンペーンシート
SD_COL = {
    'TARGETING_ID': 7,    # H列（ターゲティングID）
    'BID': 25,            # Z列（入札額）
    'CLICKS': 31,         # AF列（クリック数）
    'SPEND': 33,          # AH列（支出）
    'SALES': 34,          # AI列（売上）
    'ORDERS': 35,         # AJ列（注文数）
    'ACOS': 38,           # AM列（ACOS）
    'TARGET_ACOS': 47,    # AV列（許容ACOS）
}


# ==================== スタイル定義 ====================
class AppStyle:
    BG_PRIMARY = "#FFFFFF"
    BG_SECONDARY = "#F5F7FA"
    SURFACE = "#FFFFFF"
    SURFACE_VARIANT = "#F8FAFC"
    PRIMARY = "#6366F1"
    PRIMARY_LIGHT = "#818CF8"
    SUCCESS = "#10B981"
    SUCCESS_LIGHT = "#D1FAE5"
    ERROR = "#EF4444"
    ERROR_LIGHT = "#FEE2E2"
    SP_COLOR = "#EC4899"
    SP_BG = "#FDF2F8"
    SB_COLOR = "#14B8A6"
    SB_BG = "#F0FDFA"
    SD_COLOR = "#8B5CF6"
    SD_BG = "#F5F3FF"
    TEXT_PRIMARY = "#1E293B"
    TEXT_SECONDARY = "#64748B"
    TEXT_TERTIARY = "#94A3B8"
    TEXT_ON_PRIMARY = "#FFFFFF"
    BORDER = "#E2E8F0"

    FONT_FAMILY = "Meiryo"
    FONT_DISPLAY = (FONT_FAMILY, 20, "bold")
    FONT_HEADLINE = (FONT_FAMILY, 16, "bold")
    FONT_TITLE = (FONT_FAMILY, 14, "bold")
    FONT_BODY = (FONT_FAMILY, 11)
    FONT_BODY_BOLD = (FONT_FAMILY, 11, "bold")
    FONT_CAPTION = (FONT_FAMILY, 10)
    FONT_SMALL = (FONT_FAMILY, 9)
    FONT_EMOJI = ("Segoe UI Emoji", 24)
    FONT_EMOJI_LARGE = ("Segoe UI Emoji", 40)
    FONT_EMOJI_SMALL = ("Segoe UI Emoji", 14)


# ==================== UIコンポーネント ====================
class ModernButton(tk.Canvas):
    def __init__(self, parent, text, command, bg_color, fg_color="#FFFFFF",
                 hover_color=None, width=180, height=48, **kwargs):
        super().__init__(parent, width=width, height=height,
                        bg=parent.cget('bg'), highlightthickness=0, **kwargs)
        self.command = command
        self.bg_color = bg_color
        self.hover_color = hover_color or self._lighten_color(bg_color)
        self.fg_color = fg_color
        self.text = text
        self.width = width
        self.height = height
        self.pressed = False
        self._draw_button(self.bg_color)
        self.bind("<Enter>", self._on_enter)
        self.bind("<Leave>", self._on_leave)
        self.bind("<Button-1>", self._on_press)
        self.bind("<ButtonRelease-1>", self._on_release)

    def _lighten_color(self, color):
        r, g, b = int(color[1:3], 16), int(color[3:5], 16), int(color[5:7], 16)
        return f"#{min(255, int(r*1.15)):02x}{min(255, int(g*1.15)):02x}{min(255, int(b*1.15)):02x}"

    def _darken_color(self, color):
        r, g, b = int(color[1:3], 16), int(color[3:5], 16), int(color[5:7], 16)
        return f"#{int(r*0.85):02x}{int(g*0.85):02x}{int(b*0.85):02x}"

    def _draw_button(self, color):
        self.delete("all")
        r, w, h = 12, self.width, self.height
        self.create_arc(0, 0, r*2, r*2, start=90, extent=90, fill=color, outline=color)
        self.create_arc(w-r*2, 0, w, r*2, start=0, extent=90, fill=color, outline=color)
        self.create_arc(0, h-r*2, r*2, h, start=180, extent=90, fill=color, outline=color)
        self.create_arc(w-r*2, h-r*2, w, h, start=270, extent=90, fill=color, outline=color)
        self.create_rectangle(r, 0, w-r, h, fill=color, outline=color)
        self.create_rectangle(0, r, w, h-r, fill=color, outline=color)
        self.create_text(w//2, h//2, text=self.text, fill=self.fg_color, font=AppStyle.FONT_BODY_BOLD)

    def _on_enter(self, e): self._draw_button(self.hover_color); self.config(cursor="hand2")
    def _on_leave(self, e): self._draw_button(self.bg_color)
    def _on_press(self, e): self.pressed = True; self._draw_button(self._darken_color(self.bg_color))
    def _on_release(self, e):
        if self.pressed:
            self.pressed = False
            self._draw_button(self.hover_color)
            if self.command: self.command()

    def set_state(self, state):
        if state == "disabled":
            self.bg_color = AppStyle.TEXT_TERTIARY
            self._draw_button(self.bg_color)
            for event in ["<Enter>", "<Leave>", "<Button-1>", "<ButtonRelease-1>"]:
                self.unbind(event)
        else:
            self.bg_color = AppStyle.PRIMARY
            self._draw_button(self.bg_color)
            self.bind("<Enter>", self._on_enter)
            self.bind("<Leave>", self._on_leave)
            self.bind("<Button-1>", self._on_press)
            self.bind("<ButtonRelease-1>", self._on_release)


# ==================== メインアプリ ====================
class AmazonBidAdjusterApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Amazon広告 入札調整ツール")
        self.root.geometry("920x850")
        self.root.resizable(True, True)
        self.root.configure(bg=AppStyle.BG_SECONDARY)

        try:
            from ctypes import windll
            windll.shcore.SetProcessDpiAwareness(1)
        except: pass

        self.file_path = tk.StringVar()
        self.file_type = tk.StringVar(value="")
        self.full_path = None
        self.file_loaded = False

        self.params = {
            'click_threshold_low': tk.IntVar(value=10),
            'click_threshold_mid': tk.IntVar(value=20),
            'click_threshold_high': tk.IntVar(value=30),
            'reflect_rate_low': tk.IntVar(value=30),
            'reflect_rate_mid': tk.IntVar(value=60),
            'reflect_rate_high': tk.IntVar(value=100),
            'reduce_rate_mid': tk.IntVar(value=25),
            'reduce_rate_high': tk.IntVar(value=50),
            'max_change': tk.IntVar(value=30),
            'min_bid': tk.IntVar(value=10),
            'top_n': tk.IntVar(value=30),
            'acos_protect_min': tk.IntVar(value=30),
            'acos_protect_max': tk.IntVar(value=35),
            'new_kw_cpc_add_3plus': tk.IntVar(value=30),
            'new_kw_cpc_add_2': tk.IntVar(value=15),
            'new_kw_order1_max': tk.IntVar(value=60),
        }

        self.create_ui()

    def create_ui(self):
        container = tk.Frame(self.root, bg=AppStyle.BG_SECONDARY)
        container.pack(fill=tk.BOTH, expand=True)

        canvas = tk.Canvas(container, bg=AppStyle.BG_SECONDARY, highlightthickness=0)
        scrollbar = ttk.Scrollbar(container, orient="vertical", command=canvas.yview)

        self.main_frame = tk.Frame(canvas, bg=AppStyle.BG_SECONDARY)
        canvas.create_window((0, 0), window=self.main_frame, anchor="nw")
        self.main_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.bind_all("<MouseWheel>", lambda e: canvas.yview_scroll(int(-1*(e.delta/120)), "units"))

        content = tk.Frame(self.main_frame, bg=AppStyle.BG_SECONDARY)
        content.pack(fill=tk.BOTH, expand=True, padx=32, pady=24)

        self._create_header(content)
        self._create_drop_zone(content)
        # ファイル情報コンテナ（調整パラメータの上に配置）
        self.file_info_container = tk.Frame(content, bg=AppStyle.BG_SECONDARY)
        # 初期状態では非表示
        self._create_params_section(content)
        self._create_action_buttons(content)
        self._create_log_area(content)

        self.log("Excelファイルをドロップしてください")

    def _create_header(self, parent):
        header = tk.Frame(parent, bg=AppStyle.BG_SECONDARY)
        header.pack(fill=tk.X, pady=(0, 24))

        title_frame = tk.Frame(header, bg=AppStyle.BG_SECONDARY)
        title_frame.pack(anchor="w")

        tk.Label(title_frame, text="Amazon広告 入札調整ツール", font=AppStyle.FONT_DISPLAY, fg=AppStyle.TEXT_PRIMARY, bg=AppStyle.BG_SECONDARY).pack(side=tk.LEFT)

    def _create_drop_zone(self, parent):
        self.drop_outer = tk.Frame(parent, bg=AppStyle.BORDER, padx=2, pady=2)
        self.drop_outer.pack(fill=tk.X, pady=(0, 16))

        self.drop_zone = tk.Frame(self.drop_outer, bg=AppStyle.SURFACE, padx=32, pady=32)
        self.drop_zone.pack(fill=tk.X)

        self.drop_inner = tk.Frame(self.drop_zone, bg=AppStyle.SURFACE)
        self.drop_inner.pack(fill=tk.X, padx=2, pady=2)

        self.drop_icon = tk.Label(self.drop_inner, text="📁", font=AppStyle.FONT_EMOJI, bg=AppStyle.SURFACE)
        self.drop_icon.pack(pady=(8, 12))

        self.drop_text = tk.Label(self.drop_inner, text="Excelファイルをここにドロップ", font=AppStyle.FONT_TITLE, fg=AppStyle.TEXT_PRIMARY, bg=AppStyle.SURFACE)
        self.drop_text.pack()

        self.drop_hint = tk.Label(self.drop_inner, text="またはクリックして選択", font=AppStyle.FONT_CAPTION, fg=AppStyle.TEXT_TERTIARY, bg=AppStyle.SURFACE)
        self.drop_hint.pack(pady=(6, 8))

        for w in [self.drop_zone, self.drop_inner, self.drop_icon, self.drop_text, self.drop_hint]:
            w.drop_target_register(DND_FILES)
            w.dnd_bind('<<Drop>>', self.on_drop)
            w.bind('<Button-1>', self.on_click_select)

    def _create_params_section(self, parent):
        # パラメータセクション全体を包むフレーム
        self.params_section_frame = tk.Frame(parent, bg=AppStyle.BG_SECONDARY)
        self.params_section_frame.pack(fill=tk.X)

        section_header = tk.Frame(self.params_section_frame, bg=AppStyle.BG_SECONDARY)
        section_header.pack(fill=tk.X, pady=(8, 12))
        tk.Label(section_header, text="調整パラメータ", font=AppStyle.FONT_HEADLINE, fg=AppStyle.TEXT_PRIMARY, bg=AppStyle.BG_SECONDARY).pack(side=tk.LEFT)

        cards = tk.Frame(self.params_section_frame, bg=AppStyle.BG_SECONDARY)
        cards.pack(fill=tk.X, pady=(0, 16))

        left = tk.Frame(cards, bg=AppStyle.BG_SECONDARY)
        left.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 8))

        right = tk.Frame(cards, bg=AppStyle.BG_SECONDARY)
        right.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(8, 0))

        self._create_param_card(left, "クリック閾値", AppStyle.PRIMARY, [
            ("低", "click_threshold_low", ""),
            ("中", "click_threshold_mid", ""),
            ("高", "click_threshold_high", ""),
        ])

        self._create_param_card(left, "制限値", AppStyle.SD_COLOR, [
            ("最大変動", "max_change", "円"),
            ("最小入札", "min_bid", "円"),
            ("上位N件", "top_n", ""),
            ("ACOS下限", "acos_protect_min", "%"),
            ("ACOS上限", "acos_protect_max", "%"),
        ])

        self._create_param_card(right, "反映率", AppStyle.SB_COLOR, [
            ("低", "reflect_rate_low", "%"),
            ("中", "reflect_rate_mid", "%"),
            ("高", "reflect_rate_high", "%"),
        ])

        self._create_param_card(right, "削減率", AppStyle.SP_COLOR, [
            ("中", "reduce_rate_mid", "%"),
            ("高", "reduce_rate_high", "%"),
        ])

        self._create_param_card(right, "新規KW候補", AppStyle.PRIMARY, [
            ("3件以上+", "new_kw_cpc_add_3plus", "円"),
            ("1-2件+", "new_kw_cpc_add_2", "円"),
            ("1件上限", "new_kw_order1_max", "円"),
        ])

    def _create_param_card(self, parent, title, accent, params):
        outer = tk.Frame(parent, bg=AppStyle.BORDER, padx=1, pady=1)
        outer.pack(fill=tk.X, pady=(0, 12))

        card = tk.Frame(outer, bg=AppStyle.SURFACE, padx=16, pady=14)
        card.pack(fill=tk.X)

        header = tk.Frame(card, bg=AppStyle.SURFACE)
        header.pack(fill=tk.X, pady=(0, 12))

        accent_line = tk.Frame(header, bg=accent, width=3, height=20)
        accent_line.pack(side=tk.LEFT, padx=(0, 10))
        accent_line.pack_propagate(False)

        tk.Label(header, text=title, font=AppStyle.FONT_BODY_BOLD, fg=AppStyle.TEXT_PRIMARY, bg=AppStyle.SURFACE).pack(side=tk.LEFT)

        for label, key, suffix in params:
            row = tk.Frame(card, bg=AppStyle.SURFACE)
            row.pack(fill=tk.X, pady=3)
            tk.Label(row, text=label, font=AppStyle.FONT_CAPTION, fg=AppStyle.TEXT_SECONDARY, bg=AppStyle.SURFACE, width=10, anchor="w").pack(side=tk.LEFT)

            entry_frame = tk.Frame(row, bg=AppStyle.BORDER, padx=1, pady=1)
            entry_frame.pack(side=tk.LEFT, padx=8)
            tk.Entry(entry_frame, textvariable=self.params[key], font=AppStyle.FONT_CAPTION, bg=AppStyle.SURFACE_VARIANT, fg=AppStyle.TEXT_PRIMARY, relief="flat", width=6, justify="center").pack()

            tk.Label(row, text=suffix, font=AppStyle.FONT_SMALL, fg=AppStyle.TEXT_TERTIARY, bg=AppStyle.SURFACE).pack(side=tk.LEFT)

    def _create_action_buttons(self, parent):
        frame = tk.Frame(parent, bg=AppStyle.BG_SECONDARY)
        frame.pack(fill=tk.X, pady=(8, 16))

        self.run_button = ModernButton(frame, text="実行", command=self.run_adjustment, bg_color=AppStyle.PRIMARY, width=200, height=48)
        self.run_button.pack(side=tk.LEFT, padx=(0, 12))

        self.reset_button = ModernButton(frame, text="リセット", command=self.reset_params, bg_color=AppStyle.TEXT_TERTIARY, fg_color=AppStyle.TEXT_PRIMARY, width=140, height=48)
        self.reset_button.pack(side=tk.LEFT)

        self.progress_frame = tk.Frame(parent, bg=AppStyle.BG_SECONDARY)
        self.progress_frame.pack(fill=tk.X, pady=(0, 8))
        self.progress = ttk.Progressbar(self.progress_frame, mode='indeterminate')
        self.progress.pack(fill=tk.X)

    def _create_log_area(self, parent):
        outer = tk.Frame(parent, bg=AppStyle.BORDER, padx=1, pady=1)
        outer.pack(fill=tk.BOTH, expand=True)

        card = tk.Frame(outer, bg=AppStyle.SURFACE, padx=16, pady=14)
        card.pack(fill=tk.BOTH, expand=True)

        header = tk.Frame(card, bg=AppStyle.SURFACE)
        header.pack(fill=tk.X, pady=(0, 10))
        tk.Label(header, text="ログ", font=AppStyle.FONT_BODY_BOLD, fg=AppStyle.TEXT_PRIMARY, bg=AppStyle.SURFACE).pack(side=tk.LEFT)

        log_frame = tk.Frame(card, bg=AppStyle.BG_SECONDARY, padx=1, pady=1)
        log_frame.pack(fill=tk.BOTH, expand=True)

        # スクロールバー付きのログエリア
        log_scroll = tk.Scrollbar(log_frame)
        log_scroll.pack(side=tk.RIGHT, fill=tk.Y)

        self.log_text = tk.Text(log_frame, height=12, wrap=tk.WORD, font=AppStyle.FONT_CAPTION,
                                bg=AppStyle.SURFACE_VARIANT, fg=AppStyle.TEXT_SECONDARY,
                                relief="flat", padx=12, pady=10,
                                yscrollcommand=log_scroll.set)
        self.log_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        log_scroll.config(command=self.log_text.yview)

    def _show_file_info(self, filename, file_type):
        for w in self.file_info_container.winfo_children(): w.destroy()
        # 調整パラメータセクションの前に挿入（ドロップゾーンの後）
        self.file_info_container.pack(fill=tk.X, pady=(0, 16), before=self.params_section_frame)

        styles = {"SP": (AppStyle.SP_COLOR, AppStyle.SP_BG), "SB": (AppStyle.SB_COLOR, AppStyle.SB_BG), "SD": (AppStyle.SD_COLOR, AppStyle.SD_BG)}
        color, bg = (AppStyle.PRIMARY, AppStyle.SURFACE_VARIANT)
        for k, v in styles.items():
            if k in file_type: color, bg = v; break

        outer = tk.Frame(self.file_info_container, bg=color, padx=2, pady=2)
        outer.pack(fill=tk.X)

        card = tk.Frame(outer, bg=bg, padx=20, pady=16)
        card.pack(fill=tk.X)

        tk.Label(card, text=f"読込済: {filename}", font=AppStyle.FONT_BODY_BOLD, fg=AppStyle.SUCCESS, bg=bg).pack(side=tk.LEFT)
        tk.Label(card, text=file_type, font=AppStyle.FONT_CAPTION, fg=color, bg=bg).pack(side=tk.RIGHT)

    def _update_drop_zone_success(self):
        self.drop_outer.configure(bg=AppStyle.SUCCESS)
        bg = AppStyle.SUCCESS_LIGHT
        for w in [self.drop_zone, self.drop_inner, self.drop_icon, self.drop_text, self.drop_hint]:
            w.configure(bg=bg)
        self.drop_icon.configure(text="✓")
        self.drop_text.configure(text="ファイル読込完了", fg=AppStyle.SUCCESS)
        self.drop_hint.configure(text="別のファイルをドロップして変更")

    def on_drop(self, event):
        path = event.data
        if path.startswith('{') and path.endswith('}'): path = path[1:-1]
        self.load_file(path)

    def on_click_select(self, event):
        path = filedialog.askopenfilename(title="Excelファイルを選択", filetypes=[("Excel files", "*.xlsx")])
        if path: self.load_file(path)

    def load_file(self, path):
        self.full_path = path
        filename = os.path.basename(path)
        self.file_path.set(filename)

        file_type = self.detect_file_type(path)
        self.file_type.set(file_type)

        if file_type != "Unknown":
            self.file_loaded = True
            self._update_drop_zone_success()
            self._show_file_info(filename, file_type)
            self.log(f"\n{'='*40}")
            self.log(f"読込: {filename}")
            self.log(f"種別: {file_type}")
        else:
            self.log(f"\nエラー: 非対応ファイル - {filename}")

    def detect_file_type(self, path):
        try:
            xl = pd.ExcelFile(path)
            for sheet in xl.sheet_names:
                if 'スポンサープロダクト' in sheet: return "SP"
                if 'スポンサーブランド' in sheet: return "SB"
                if 'スポンサーディスプレイ' in sheet: return "SD"

            name = os.path.basename(path).upper()
            if 'SP' in name: return "SP"
            if 'SB' in name: return "SB"
            if 'SD' in name: return "SD"
            return "Unknown"
        except:
            return "Unknown"

    def log(self, msg):
        self.log_text.insert(tk.END, msg + "\n")
        self.log_text.see(tk.END)
        self.root.update()

    def reset_params(self):
        defaults = {
            'click_threshold_low': 10, 'click_threshold_mid': 20, 'click_threshold_high': 30,
            'reflect_rate_low': 30, 'reflect_rate_mid': 60, 'reflect_rate_high': 100,
            'reduce_rate_mid': 25, 'reduce_rate_high': 50,
            'max_change': 30, 'min_bid': 10, 'top_n': 30,
            'acos_protect_min': 30, 'acos_protect_max': 35,
            'new_kw_cpc_add_3plus': 30, 'new_kw_cpc_add_2': 15, 'new_kw_order1_max': 60,
        }
        for k, v in defaults.items(): self.params[k].set(v)
        self.log("\nパラメータをリセットしました")

    def run_adjustment(self):
        if not self.full_path:
            messagebox.showerror("エラー", "ファイルを選択してください")
            return

        self.run_button.set_state("disabled")
        self.progress.start()
        threading.Thread(target=self.execute_adjustment).start()

    def execute_adjustment(self):
        try:
            file_type = self.file_type.get()
            if "SP" in file_type:
                self.run_sp_sb_adjustment(SP_COL, SP_SEARCH_COL, "SP")
            elif "SB" in file_type:
                self.run_sp_sb_adjustment(SB_COL, SB_SEARCH_COL, "SB")
            elif "SD" in file_type:
                self.run_sd_adjustment()
        except Exception as e:
            self.log(f"\nエラー: {str(e)}")
            messagebox.showerror("エラー", str(e))
        finally:
            self.progress.stop()
            self.run_button.set_state("normal")

    def safe_float(self, value, default=0.0):
        if pd.isna(value) or value == '' or value == '-': return default
        try:
            if isinstance(value, str):
                value = value.replace('%', '').replace(',', '').replace('yen', '').strip()
            return float(value)
        except: return default

    def normalize_keyword(self, value):
        text = str(value).strip().lower()
        text = text.replace("\u3000", " ")
        return " ".join(text.split())

    def is_manual_campaign(self, name):
        text = str(name).strip()
        if "マニュアル" in text:
            return True
        return bool(re.search(r"(M\d*$|ｍ\d*$)", text))

    def extract_asins(self, value):
        if pd.isna(value) or value == '':
            return set()
        text = str(value).upper()
        return set(re.findall(r'[A-Z0-9]{10}', text))

    def is_asin(self, value):
        text = str(value).strip().upper()
        return bool(re.fullmatch(r'[A-Z0-9]{10}', text))

    def apply_limit(self, current, new_bid, max_change):
        diff = new_bid - current
        if diff > max_change: return current + max_change, True
        if diff < -max_change: return current - max_change, True
        return new_bid, False

    def run_sp_sb_adjustment(self, col, search_col, ad_type):
        self.log(f"\n{'='*40}")
        self.log(f"{ad_type}の調整を開始")

        df_camp = pd.read_excel(self.full_path, sheet_name=0)
        df_search = pd.read_excel(self.full_path, sheet_name=1)

        self.log(f"キャンペーン: {len(df_camp)}行")
        self.log(f"検索ワード: {len(df_search)}行")

        p = {k: v.get() for k, v in self.params.items()}

        search_cv = {}
        target_acos_map = {}
        original_bid_map = {}
        search_term_data = []
        manual_keywords = set()
        product_target_asins = set()
        manual_keyword_campaigns = {}
        product_target_campaigns = {}
        portfolio_has_manual = set()
        portfolio_has_product = set()
        sp_manual_kw_targets = set()
        sp_product_asins_by_campaign = set()
        sp_manual_campaign_ids = {}
        sp_product_campaign_ids = {}
        campaign_id_to_name = {}
        campaign_id_to_ad_groups = {}
        campaign_name_info_idx = None
        if 'キャンペーン名（情報提供のみ）' in df_camp.columns:
            campaign_name_info_idx = df_camp.columns.get_loc('キャンペーン名（情報提供のみ）')

        def add_campaign(campaign_map, key, campaign_id):
            if not campaign_id:
                return
            if key not in campaign_map:
                campaign_map[key] = set()
            campaign_map[key].add(campaign_id)

        # まず、キャンペーンシートからtarget_acos_mapとoriginal_bid_mapを構築
        # （検索ワードシートには許容ACOSが存在しない）
        # 同時に既存キーワード/ターゲティングを収集
        for _, row in df_camp.iterrows():
            kw_id = str(int(row.iloc[col['KEYWORD_ID']])) if pd.notna(row.iloc[col['KEYWORD_ID']]) else ""
            tg_id = str(int(row.iloc[col['TARGETING_ID']])) if pd.notna(row.iloc[col['TARGETING_ID']]) else ""
            id_key = kw_id if kw_id and kw_id != 'nan' else tg_id

            if not id_key or id_key == 'nan':
                continue
            campaign_id = str(int(row.iloc[col['CAMPAIGN_ID']])) if pd.notna(row.iloc[col['CAMPAIGN_ID']]) else ""
            if 'AD_GROUP_ID' in col:
                ad_group_id_camp = str(int(row.iloc[col['AD_GROUP_ID']])) if pd.notna(row.iloc[col['AD_GROUP_ID']]) else ""
                if campaign_id and ad_group_id_camp:
                    if campaign_id not in campaign_id_to_ad_groups:
                        campaign_id_to_ad_groups[campaign_id] = set()
                    campaign_id_to_ad_groups[campaign_id].add(ad_group_id_camp)
            if campaign_name_info_idx is not None:
                campaign_name_info = str(row.iloc[campaign_name_info_idx]).strip() if pd.notna(row.iloc[campaign_name_info_idx]) else ""
            elif 'CAMPAIGN_NAME_INFO' in col:
                campaign_name_info = str(row.iloc[col['CAMPAIGN_NAME_INFO']]).strip() if pd.notna(row.iloc[col['CAMPAIGN_NAME_INFO']]) else ""
            else:
                campaign_name_info = ""
            if campaign_id and campaign_id not in campaign_id_to_name:
                campaign_id_to_name[campaign_id] = campaign_name_info

            target_acos_val = self.safe_float(row.iloc[col['TARGET_ACOS']], 0)
            if target_acos_val > 0 and id_key not in target_acos_map:
                target_acos_map[id_key] = target_acos_val

            # 元の入札額を保存
            bid_val = self.safe_float(row.iloc[col['BID']], 0)
            if bid_val > 0 and id_key not in original_bid_map:
                original_bid_map[id_key] = bid_val

            # 既存キーワード/ターゲティングを収集（ポートフォリオ単位）
            portfolio = str(row.iloc[col['PORTFOLIO']]) if pd.notna(row.iloc[col['PORTFOLIO']]) else ""

            # キーワードテキスト
            if 'KEYWORD_TEXT' in col:
                kw_text = self.normalize_keyword(row.iloc[col['KEYWORD_TEXT']]) if pd.notna(row.iloc[col['KEYWORD_TEXT']]) else ""
                if kw_text and kw_text != 'nan':
                    key = (portfolio, kw_text)
                    is_manual = self.is_manual_campaign(campaign_name_info)
                    if ad_type == "SB" and not is_manual:
                        pass
                    else:
                        manual_keywords.add(key)
                        portfolio_has_manual.add(portfolio)
                        add_campaign(manual_keyword_campaigns, key, campaign_id)
                    if ad_type == "SP" and is_manual:
                        match_type = str(row.iloc[col['MATCH_TYPE']]).strip() if pd.notna(row.iloc[col['MATCH_TYPE']]) else ""
                        entity = str(row.iloc[col['ENTITY']]).strip() if pd.notna(row.iloc[col['ENTITY']]) else ""
                        if entity == "キーワード" and match_type and kw_text:
                            sp_manual_kw_targets.add((portfolio, kw_text, match_type))

            # 商品ターゲティング式
            if 'PRODUCT_TARGETING' in col:
                pt_expr = str(row.iloc[col['PRODUCT_TARGETING']]).strip() if pd.notna(row.iloc[col['PRODUCT_TARGETING']]) else ""
                if pt_expr and pt_expr != 'nan':
                    portfolio_has_product.add(portfolio)
                    for asin in self.extract_asins(pt_expr):
                        key = (portfolio, asin)
                        is_product = "商品" in campaign_name_info
                        if ad_type == "SB" and not is_product:
                            pass
                        else:
                            product_target_asins.add(key)
                            add_campaign(product_target_campaigns, key, campaign_id)
                        if ad_type == "SP" and is_product:
                            sp_product_asins_by_campaign.add((portfolio, asin))

            if ad_type == "SP" and campaign_name_info:
                if self.is_manual_campaign(campaign_name_info):
                    add_campaign(sp_manual_campaign_ids, portfolio, campaign_id)
                if "商品" in campaign_name_info:
                    add_campaign(sp_product_campaign_ids, portfolio, campaign_id)

        self.log(f"既存キーワードターゲティング: {len(manual_keywords)}件")
        self.log(f"既存商品ターゲティング(ASIN): {len(product_target_asins)}件")

        # 検索ワードシートからCV情報を取得
        for _, row in df_search.iterrows():
            kw_id = str(int(row.iloc[search_col['KEYWORD_ID']])) if pd.notna(row.iloc[search_col['KEYWORD_ID']]) else ""
            tg_id = str(int(row.iloc[search_col['TARGETING_ID']])) if pd.notna(row.iloc[search_col['TARGETING_ID']]) else ""
            id_key = kw_id if kw_id and kw_id != 'nan' else tg_id

            if not id_key or id_key == 'nan': continue

            orders = self.safe_float(row.iloc[search_col['ORDERS']], 0)
            clicks = self.safe_float(row.iloc[search_col['CLICKS']], 0)

            if id_key not in search_cv:
                search_cv[id_key] = {'has_cv': False, 'total_clicks': 0, 'cv_terms': 0}

            search_cv[id_key]['total_clicks'] += clicks
            if orders > 0:
                search_cv[id_key]['has_cv'] = True
                search_cv[id_key]['cv_terms'] += 1

            portfolio = str(row.iloc[search_col['PORTFOLIO']]) if pd.notna(row.iloc[search_col['PORTFOLIO']]) else ""
            search_term = str(row.iloc[search_col['SEARCH_TERM']]) if pd.notna(row.iloc[search_col['SEARCH_TERM']]) else ""
            keyword_text = str(row.iloc[search_col['KEYWORD_TEXT']]).strip() if pd.notna(row.iloc[search_col['KEYWORD_TEXT']]) else ""
            match_type = str(row.iloc[search_col['MATCH_TYPE']]).strip() if pd.notna(row.iloc[search_col['MATCH_TYPE']]) else ""
            spend = self.safe_float(row.iloc[search_col['SPEND']], 0)
            sales = self.safe_float(row.iloc[search_col['SALES']], 0)

            # キャンペーンID・グループID・名前を取得
            campaign_id = str(int(row.iloc[search_col['CAMPAIGN_ID']])) if pd.notna(row.iloc[search_col['CAMPAIGN_ID']]) else ""
            ad_group_id = str(int(row.iloc[search_col['AD_GROUP_ID']])) if pd.notna(row.iloc[search_col['AD_GROUP_ID']]) else ""
            campaign_name = str(row.iloc[search_col['CAMPAIGN_NAME']]) if pd.notna(row.iloc[search_col['CAMPAIGN_NAME']]) else ""
            ad_group_name = str(row.iloc[search_col['AD_GROUP_NAME']]) if pd.notna(row.iloc[search_col['AD_GROUP_NAME']]) else ""
            if campaign_id and ad_group_id:
                if campaign_id not in campaign_id_to_ad_groups:
                    campaign_id_to_ad_groups[campaign_id] = set()
                campaign_id_to_ad_groups[campaign_id].add(ad_group_id)

            if search_term and search_term != 'nan':
                search_term_data.append({
                    'search_term': search_term, 'portfolio': portfolio, 'id_key': id_key,
                    'campaign_id': campaign_id, 'ad_group_id': ad_group_id,
                    'campaign_name': campaign_name, 'ad_group_name': ad_group_name,
                    'clicks': clicks, 'spend': spend, 'sales': sales, 'orders': orders,
                    'keyword_text': keyword_text, 'match_type': match_type
                })

        df_sorted = df_camp.copy()
        df_sorted['_sales'] = df_sorted.iloc[:, col['SALES']].apply(lambda x: self.safe_float(x, 0))
        top_n_idx = set(df_sorted.sort_values('_sales', ascending=False).head(p['top_n']).index)

        new_bids, original_bids, reasons = [], [], []
        stats = {}

        for idx, row in df_camp.iterrows():
            kw_id = str(int(row.iloc[col['KEYWORD_ID']])) if pd.notna(row.iloc[col['KEYWORD_ID']]) else ""
            tg_id = str(int(row.iloc[col['TARGETING_ID']])) if pd.notna(row.iloc[col['TARGETING_ID']]) else ""
            id_key = kw_id if kw_id and kw_id != 'nan' else tg_id

            current = self.safe_float(row.iloc[col['BID']], 0)
            clicks = self.safe_float(row.iloc[col['CLICKS']], 0)
            acos = self.safe_float(row.iloc[col['ACOS']], 0)
            target_acos = self.safe_float(row.iloc[col['TARGET_ACOS']], 0)

            original_bids.append(current)

            if target_acos == 0 and id_key in target_acos_map:
                target_acos = target_acos_map[id_key]

            cv_info = search_cv.get(id_key, {'has_cv': False, 'cv_terms': 0})
            has_cv = cv_info['has_cv']
            cv_terms = cv_info['cv_terms']

            reason, new_bid = self._calculate_bid(
                current, clicks, acos, target_acos, has_cv, cv_terms, idx in top_n_idx, p
            )

            new_bids.append(new_bid)
            reasons.append(reason)
            # 理由をそのままキーとして集計
            stats[reason] = stats.get(reason, 0) + 1

        # カテゴリ別に整理してログ出力
        self._log_organized_summary(stats)

        new_kw_candidates, skipped_existing = self._find_new_keywords(
            search_term_data,
            target_acos_map,
            original_bid_map,
            p,
            portfolio_has_manual,
            portfolio_has_product,
            manual_keywords,
            product_target_asins,
            manual_keyword_campaigns,
            product_target_campaigns,
            ad_type=ad_type,
            sp_manual_kw_targets=sp_manual_kw_targets,
            sp_product_asins_by_campaign=sp_product_asins_by_campaign,
            sp_manual_campaign_ids=sp_manual_campaign_ids,
            sp_product_campaign_ids=sp_product_campaign_ids,
            campaign_id_to_name=campaign_id_to_name,
            campaign_id_to_ad_groups=campaign_id_to_ad_groups,
        )
        exclude_candidates = self._find_exclude_keywords(search_term_data, p)

        self.log(f"\n【キーワード候補】")
        self.log(f"  新規キーワード候補: {len(new_kw_candidates)}件")
        self.log(f"  既存キーワード除外: {skipped_existing}件")
        self.log(f"  除外キーワード候補: {len(exclude_candidates)}件")

        output = self.full_path.replace('.xlsx', '_adjusted.xlsx')
        wb = load_workbook(self.full_path)
        ws = wb.worksheets[0]

        out_col = 55
        ws.cell(row=1, column=out_col, value="元の入札額")
        ws.cell(row=1, column=out_col+1, value="新入札額")
        ws.cell(row=1, column=out_col+2, value="理由")

        for i, (o, n, r) in enumerate(zip(original_bids, new_bids, reasons)):
            ws.cell(row=i+2, column=out_col, value=o)
            ws.cell(row=i+2, column=out_col+1, value=n)
            ws.cell(row=i+2, column=out_col+2, value=r)

        # 既存の候補シートを削除（重複防止）
        sheets_to_remove = ['新規キーワード候補', '除外候補', '除外キーワード候補']
        for sheet_name in sheets_to_remove:
            if sheet_name in wb.sheetnames:
                del wb[sheet_name]

        if new_kw_candidates:
            ws_new = wb.create_sheet("新規キーワード候補")
            ws_new.append(list(new_kw_candidates[0].keys()))
            for c in new_kw_candidates: ws_new.append(list(c.values()))
            # 追加列を薄い色でハイライト
            highlight_fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
            header = [cell.value for cell in ws_new[1]]
            try:
                id_idx = header.index("同一ポートフォリオ内ターゲティングキャンペーンID") + 1
                name_idx = header.index("同一ポートフォリオ内ターゲティングキャンペーン名") + 1
                add_kw_idx = header.index("追加キーワード(完全一致)") + 1
                add_mt_idx = header.index("追加マッチタイプ") + 1
                click_idx = header.index("クリック数") + 1
                if not (add_kw_idx == click_idx - 2 and add_mt_idx == click_idx - 1):
                    kw_vals = [ws_new.cell(row=r, column=add_kw_idx).value for r in range(1, ws_new.max_row + 1)]
                    mt_vals = [ws_new.cell(row=r, column=add_mt_idx).value for r in range(1, ws_new.max_row + 1)]
                    ws_new.insert_cols(click_idx, amount=2)
                    for row in range(1, ws_new.max_row + 1):
                        ws_new.cell(row=row, column=click_idx).value = kw_vals[row - 1]
                        ws_new.cell(row=row, column=click_idx + 1).value = mt_vals[row - 1]
                    if add_kw_idx >= click_idx:
                        add_kw_idx += 2
                    if add_mt_idx >= click_idx:
                        add_mt_idx += 2
                    for idx in sorted([add_kw_idx, add_mt_idx], reverse=True):
                        ws_new.delete_cols(idx)
                    header = [cell.value for cell in ws_new[1]]
                    add_kw_idx = header.index("追加キーワード(完全一致)") + 1
                    add_mt_idx = header.index("追加マッチタイプ") + 1
                for row in ws_new.iter_rows(min_row=1, max_row=ws_new.max_row, min_col=id_idx, max_col=name_idx):
                    for cell in row:
                        cell.fill = highlight_fill
                for row in ws_new.iter_rows(min_row=1, max_row=ws_new.max_row, min_col=add_kw_idx, max_col=add_kw_idx):
                    for cell in row:
                        cell.fill = highlight_fill
                for row in ws_new.iter_rows(min_row=1, max_row=ws_new.max_row, min_col=add_mt_idx, max_col=add_mt_idx):
                    for cell in row:
                        cell.fill = highlight_fill
            except ValueError:
                pass

        if exclude_candidates:
            ws_ex = wb.create_sheet("除外候補")
            ws_ex.append(list(exclude_candidates[0].keys()))
            for c in exclude_candidates: ws_ex.append(list(c.values()))

        wb.save(output)
        self.log(f"\n完了: {os.path.basename(output)}")
        messagebox.showinfo("完了", f"処理が完了しました\n\n{output}")

    def _calculate_bid(self, current, clicks, acos, target_acos, has_cv, cv_terms, is_top, p):
        min_bid = p['min_bid']
        max_change = p['max_change']

        # 1. 許容ACOS未設定
        if target_acos == 0:
            return "許容ACOS未設定", current

        # 2. クリック数10以下
        if clicks <= p['click_threshold_low']:
            return f"クリック{p['click_threshold_low']}以下（データ不足）", current

        # 3. ACOS=0（売上なし）+ クリック11〜20 → 10%削減（制限なし）
        if acos == 0 and p['click_threshold_low'] < clicks <= p['click_threshold_mid']:
            new = max(min_bid, round(current * 0.90))
            return f"売上なし+クリック{p['click_threshold_low']+1}-{p['click_threshold_mid']}→10%削減", new

        # 4. ACOS=0（売上なし）+ クリック21〜29 → 25%削減（制限なし）
        if acos == 0 and p['click_threshold_mid'] < clicks < p['click_threshold_high']:
            new = max(min_bid, round(current * 0.75))
            return f"売上なし+クリック{p['click_threshold_mid']+1}-{p['click_threshold_high']-1}→25%削減", new

        # 5. ACOS=0（売上なし）+ クリック30以上 → 50%削減（制限なし）
        if acos == 0 and clicks >= p['click_threshold_high']:
            new = max(min_bid, round(current * 0.50))
            return f"売上なし+クリック{p['click_threshold_high']}以上→50%削減", new

        # 6. 売上上位N位 + ACOS 30%〜35% → 現状維持
        acos_min = p['acos_protect_min'] / 100
        acos_max = p['acos_protect_max'] / 100
        if is_top and acos_min <= acos <= acos_max:
            return f"上位商品・ACOS適正({acos*100:.1f}%)、現状維持", current

        if acos == 0:
            return "ACOS計算不可", current

        # 基本計算
        ratio = target_acos / acos
        calculated = current * ratio
        adjustment = calculated - current

        # 7. クリック11〜20件（売上あり）→ 30%反映
        if p['click_threshold_low'] < clicks <= p['click_threshold_mid']:
            new = current + (adjustment * p['reflect_rate_low'] / 100)
            new, limited = self.apply_limit(current, new, max_change)
            new = max(min_bid, round(new))
            direction = "↑" if new > current else "↓" if new < current else "→"
            limit_note = "（±30円制限）" if limited else ""
            return f"クリック{p['click_threshold_low']+1}-{p['click_threshold_mid']}（{p['reflect_rate_low']}%反映）、入札{direction}{limit_note}", new

        # 8. クリック21〜29件（売上あり）→ 60%反映
        if p['click_threshold_mid'] < clicks < p['click_threshold_high']:
            new = current + (adjustment * p['reflect_rate_mid'] / 100)
            new, limited = self.apply_limit(current, new, max_change)
            new = max(min_bid, round(new))
            direction = "↑" if new > current else "↓" if new < current else "→"
            limit_note = "（±30円制限）" if limited else ""
            return f"クリック{p['click_threshold_mid']+1}-{p['click_threshold_high']-1}（{p['reflect_rate_mid']}%反映）、入札{direction}{limit_note}", new

        # 9. 検索語CV無し + クリック30以上 → 50%削減
        if not has_cv and clicks >= p['click_threshold_high']:
            new = current * 0.50
            new, limited = self.apply_limit(current, new, max_change)
            new = max(min_bid, round(new))
            limit_note = "（±30円制限）" if limited else ""
            return f"検索語CV無し+クリック{p['click_threshold_high']}以上→50%削減{limit_note}", new

        # 10. 通常調整（クリック30以上・CV有り）
        if clicks >= p['click_threshold_high']:
            new, limited = self.apply_limit(current, calculated, max_change)
            new = max(min_bid, round(new))
            direction = "↑" if new > current else "↓" if new < current else "→"
            limit_note = "（±30円制限）" if limited else ""
            return f"通常調整（クリック{p['click_threshold_high']}以上）、入札{direction}{limit_note}", new

        # その他
        return "その他", current

    def _find_new_keywords(self, data, target_acos_map, original_bid_map, p, *args, **kwargs):
        ad_type = kwargs.get('ad_type')
        sp_manual_kw_targets = kwargs.get('sp_manual_kw_targets', set())
        sp_product_asins_by_campaign = kwargs.get('sp_product_asins_by_campaign', set())
        sp_manual_campaign_ids = kwargs.get('sp_manual_campaign_ids', {})
        sp_product_campaign_ids = kwargs.get('sp_product_campaign_ids', {})
        campaign_id_to_name = kwargs.get('campaign_id_to_name', {})
        campaign_id_to_ad_groups = kwargs.get('campaign_id_to_ad_groups', {})
        if len(args) == 1:
            existing_keywords = args[0]
            manual_keywords = set()
            for term, portfolio in existing_keywords:
                manual_keywords.add((portfolio, str(term).strip().lower()))
            portfolio_has_manual = {portfolio for _, portfolio in existing_keywords}
            portfolio_has_product = set()
            product_target_asins = set()
            manual_keyword_campaigns = {}
            product_target_campaigns = {}
        else:
            (portfolio_has_manual, portfolio_has_product,
             manual_keywords, product_target_asins,
             manual_keyword_campaigns, product_target_campaigns) = args
        portfolio_terms = {}
        for item in data:
            key = (item['search_term'], item['portfolio'])
            if key not in portfolio_terms:
                portfolio_terms[key] = []
            portfolio_terms[key].append(item)

        candidates = []
        skipped_existing = 0  # 既存キーワードとしてスキップした件数
        sp_keyword_total = 0
        sp_asin_total = 0
        sp_keyword_existing = 0
        sp_asin_existing = 0
        sp_keyword_output = 0
        sp_asin_output = 0
        sp_other_skipped = 0
        candidate_keys = set()

        for (term, portfolio), items in portfolio_terms.items():
            # 既存キーワードチェック（大文字小文字を区別しない）
            term_normalized = self.normalize_keyword(term)
            term_asin = term.strip().upper()
            is_asin = self.is_asin(term_asin)
            if ad_type == "SP":
                item = items[0]
                keyword_text_field = self.normalize_keyword(item.get('keyword_text', ''))
                match_type = str(item.get('match_type', '')).strip()
                campaign_name = str(item.get('campaign_name', '')).strip()
                is_auto_campaign = "オート" in campaign_name
                is_keyword_case = keyword_text_field != '' and match_type != ''
                is_asin_case = is_asin and keyword_text_field == '' and match_type == ''
                if is_keyword_case:
                    sp_keyword_total += 1
                    if (portfolio, term_normalized, "完全一致") in sp_manual_kw_targets:
                        sp_keyword_existing += 1
                        skipped_existing += 1
                        continue
                elif is_asin_case:
                    sp_asin_total += 1
                    if (portfolio, term_asin) in sp_product_asins_by_campaign:
                        sp_asin_existing += 1
                        skipped_existing += 1
                        continue
                elif is_auto_campaign and not is_asin:
                    sp_keyword_total += 1
                    if (portfolio, term_normalized, "完全一致") in sp_manual_kw_targets:
                        sp_keyword_existing += 1
                        skipped_existing += 1
                        continue
                else:
                    sp_other_skipped += 1
                    continue
            else:
                has_targeting = (portfolio in portfolio_has_manual) or (portfolio in portfolio_has_product)
                if has_targeting:
                    if is_asin:
                        if (portfolio, term_asin) in product_target_asins:
                            skipped_existing += 1
                            continue
                    else:
                        if (portfolio, term_normalized) in manual_keywords:
                            skipped_existing += 1
                            continue

            if len(items) == 1 and items[0]['orders'] > 0:
                item = items[0]
                if ad_type == "SP":
                    if is_asin:
                        sp_asin_output += 1
                    else:
                        sp_keyword_output += 1
                orders = int(item['orders'])
                clicks = int(item['clicks'])
                if orders < 2 and clicks < 30:
                    continue
                cpc = item['spend'] / item['clicks'] if item['clicks'] > 0 else 0
                # ACOSは小数で計算（例: 0.282 = 28.2%）
                acos = item['spend'] / item['sales'] if item['sales'] > 0 else 0
                cvr = item['orders'] / item['clicks'] if item['clicks'] > 0 else 0
                # target_acos_mapの値はキャンペーンシートから取得（小数形式: 0.30 = 30%）
                target = target_acos_map.get(item['id_key'], 0)
                # 元の入札額を取得
                original_bid = original_bid_map.get(item['id_key'], 0)

                rec_bid = "N/A"
                target_str = "N/A"
                action = "N/A"
                orders = int(item['orders'])
                # 両方小数で計算（target/acos比率）
                if target > 0 and acos > 0:
                    raw_rec = round(cpc * (target / acos))
                    rec = max(raw_rec, p['min_bid'])
                    # 入札上限ルール: 注文数に応じてCPC上乗せ上限を変える
                    cpc_rounded = round(cpc)
                    cpc_add_3plus = p.get('new_kw_cpc_add_3plus', 30)
                    cpc_add_2 = p.get('new_kw_cpc_add_2', 15)
                    order1_max = p.get('new_kw_order1_max', 60)

                    # 方向判定（CPC基準）
                    if rec > cpc_rounded:
                        direction = "↑"
                    elif rec < cpc_rounded:
                        direction = "↓"
                    else:
                        direction = "→"

                    if orders >= 3:
                        # 注文3件以上: CPC + 30円まで
                        if rec > cpc_rounded + cpc_add_3plus:
                            rec = cpc_rounded + cpc_add_3plus
                            action = f"注文{orders}件・{direction}上限+{cpc_add_3plus}円"
                        else:
                            diff = rec - cpc_rounded
                            if diff > 0:
                                action = f"注文{orders}件・{direction}+{diff}円"
                            elif diff < 0:
                                action = f"注文{orders}件・{direction}{diff}円"
                            else:
                                action = f"注文{orders}件・維持"
                    elif orders == 2:
                        # 注文2件: CPC + 15円まで
                        if rec > cpc_rounded + cpc_add_2:
                            rec = cpc_rounded + cpc_add_2
                            action = f"注文2件・{direction}上限+{cpc_add_2}円"
                        else:
                            diff = rec - cpc_rounded
                            if diff > 0:
                                action = f"注文2件・{direction}+{diff}円"
                            elif diff < 0:
                                action = f"注文2件・{direction}{diff}円"
                            else:
                                action = f"注文2件・維持"
                    else:
                        # 注文1件: CPC + 15円まで、かつ最大60円
                        hit_cpc_limit = False
                        hit_max_limit = False
                        if rec > cpc_rounded + cpc_add_2:
                            rec = cpc_rounded + cpc_add_2
                            hit_cpc_limit = True
                        if rec > order1_max:
                            rec = order1_max
                            hit_max_limit = True

                        diff = rec - cpc_rounded
                        if hit_max_limit:
                            action = f"注文1件・{direction}上限{order1_max}円"
                        elif hit_cpc_limit:
                            action = f"注文1件・{direction}上限+{cpc_add_2}円"
                        elif diff > 0:
                            action = f"注文1件・{direction}+{diff}円"
                        elif diff < 0:
                            action = f"注文1件・{direction}{diff}円"
                        else:
                            action = f"注文1件・維持"

                    rec_bid = f"{rec}円"
                    # キャンペーンシートの許容ACOSは小数形式（0.30 = 30%）
                    target_str = f"{target*100:.0f}%"

                # ACOSは計算値なので常に小数
                acos_str = f"{acos*100:.1f}%"
                related_campaign_ids = ""
                related_ad_group_ids = ""
                match_type_value = str(item.get('match_type', '')).strip()
                add_match_type = "完全一致" if ad_type == "SP" and keyword_text_field != '' and match_type_value != '' else ""
                if ad_type == "SP":
                    keyword_text_field = self.normalize_keyword(item.get('keyword_text', ''))
                    if is_asin and keyword_text_field == '' and match_type_value == '':
                        related = sp_product_campaign_ids.get(portfolio, set())
                    else:
                        related = sp_manual_campaign_ids.get(portfolio, set())
                else:
                    if is_asin:
                        related = product_target_campaigns.get((portfolio, term_asin), set())
                    else:
                        related = manual_keyword_campaigns.get((portfolio, term_normalized), set())
                related_campaign_names = ""
                if related:
                    related_campaign_ids = ", ".join(sorted(related))
                    related_campaign_names = ", ".join(
                        [campaign_id_to_name.get(cid, "") for cid in sorted(related) if campaign_id_to_name.get(cid, "")]
                    )
                    if ad_type == "SP":
                        related_ad_group_ids = ", ".join(
                            sorted({gid for cid in related for gid in campaign_id_to_ad_groups.get(cid, set())})
                        )
                campaign_name_for_id = campaign_id_to_name.get(item.get('campaign_id', ''), item.get('campaign_name', ''))
                is_asin_case = is_asin and self.normalize_keyword(item.get('keyword_text', '')) == '' and match_type_value == ''
                asin_note = f"ASIN {term} を商品ターゲティングに追加" if is_asin_case else ""

                candidate_key = (portfolio, term_normalized, is_asin)
                if candidate_key in candidate_keys:
                    continue
                candidate_keys.add(candidate_key)

                target_type = "ASIN" if is_asin_case else "キーワード"
                candidates.append({
                    'カスタマー検索語': term,
                    'ポートフォリオ名': portfolio,
                    'キャンペーンID': item.get('campaign_id', ''),
                    'キャンペーン名': campaign_name_for_id,
                    '追加対象種別': target_type,
                    '同一ポートフォリオ内ターゲティングキャンペーンID': related_campaign_ids,
                    '同一ポートフォリオ内ターゲティング広告グループID': related_ad_group_ids,
                    '同一ポートフォリオ内ターゲティングキャンペーン名': related_campaign_names,
                    '広告グループID': item.get('ad_group_id', ''),
                    '広告グループ名': item.get('ad_group_name', ''),
                    '元ターゲティングID': item['id_key'],
                    '追加キーワード(完全一致)': term,
                    '追加マッチタイプ': add_match_type or match_type_value,
                    'クリック数': int(item['clicks']),
                    '売上': round(item['sales']), '注文数': int(item['orders']),
                    'CPC': f"{cpc:.1f}円", 'ACOS': acos_str,
                    'コンバージョン率': f"{cvr*100:.2f}%",
                    '推奨入札単価': rec_bid, '許容ACOS': target_str,
                    'ASIN設定メモ': asin_note,
                    '推奨アクション': action
                })

        if ad_type == "SP":
            self.log(
                "SP新規候補判定: "
                f"KW総数={sp_keyword_total}, KW設定済み={sp_keyword_existing}, KW出力={sp_keyword_output}, "
                f"ASIN総数={sp_asin_total}, ASIN設定済み={sp_asin_existing}, ASIN出力={sp_asin_output}, "
                f"その他除外={sp_other_skipped}"
            )
        sorted_candidates = sorted(candidates, key=lambda x: x['クリック数'], reverse=True)
        return sorted_candidates, skipped_existing

    def _find_exclude_keywords(self, data, p):
        portfolio_terms = {}
        for item in data:
            key = (item['search_term'], item['portfolio'])
            if key not in portfolio_terms:
                portfolio_terms[key] = []
            portfolio_terms[key].append(item)

        candidates = []
        for (term, portfolio), items in portfolio_terms.items():
            total_clicks = sum(i['clicks'] for i in items)
            total_orders = sum(i['orders'] for i in items)

            if total_clicks >= p['click_threshold_high'] and total_orders == 0:
                total_spend = sum(i['spend'] for i in items)
                cpc = total_spend / total_clicks if total_clicks > 0 else 0
                related_ids = list(set([i['id_key'] for i in items if i['id_key'] and i['id_key'] != 'nan']))

                # 最初のアイテムからキャンペーン・グループ情報を取得
                first_item = items[0]
                campaign_id = first_item.get('campaign_id', '')
                ad_group_id = first_item.get('ad_group_id', '')
                campaign_name = first_item.get('campaign_name', '')
                ad_group_name = first_item.get('ad_group_name', '')

                candidates.append({
                    'カスタマー検索語': term, 'ポートフォリオ名': portfolio,
                    'キャンペーンID': campaign_id,
                    '広告グループID': ad_group_id,
                    'キャンペーン名': campaign_name,
                    '広告グループ名': ad_group_name,
                    '関連ターゲティング数': len(related_ids),
                    '関連ターゲティングID': ', '.join(related_ids[:5]),
                    '合計クリック数': int(total_clicks), '合計支出': f"{total_spend:.0f}円",
                    'CPC': f"{cpc:.1f}円", 'ACOS': 'N/A', 'コンバージョン率': "0.00%",
                    '推奨アクション': '除外キーワード追加検討'
                })

        return sorted(candidates, key=lambda x: x['合計クリック数'], reverse=True)

    def _log_organized_summary(self, stats):
        """調整結果をカテゴリ別に整理してログ出力"""
        self.log("\n" + "="*50)
        self.log("調整結果サマリー")
        self.log("="*50)

        # カテゴリ定義（表示順序）
        categories = {
            'データ不足': [],
            '売上なし（ACOS=0）': [],
            '売上あり・通常調整': [],
            '売上あり・部分反映': [],
            '検索語CV関連': [],
            '上位商品保護': [],
            '設定・その他': [],
        }

        # 理由をカテゴリに分類
        for reason, count in stats.items():
            if 'データ不足' in reason or 'クリック10以下' in reason:
                categories['データ不足'].append((reason, count))
            elif '売上なし' in reason or 'ACOS=0' in reason:
                categories['売上なし（ACOS=0）'].append((reason, count))
            elif '通常調整' in reason:
                categories['売上あり・通常調整'].append((reason, count))
            elif '反映' in reason:
                categories['売上あり・部分反映'].append((reason, count))
            elif '検索語CV' in reason:
                categories['検索語CV関連'].append((reason, count))
            elif '上位商品' in reason or '現状維持' in reason:
                categories['上位商品保護'].append((reason, count))
            else:
                categories['設定・その他'].append((reason, count))

        # 合計を計算
        total = sum(stats.values())

        # カテゴリ別に出力
        for cat_name, items in categories.items():
            if items:
                cat_total = sum(c for _, c in items)
                self.log(f"\n【{cat_name}】 計{cat_total}件")

                # 入札↑、入札→、入札↓の順でソート
                def sort_key(item):
                    reason = item[0]
                    if '↑' in reason:
                        return (0, -item[1])
                    elif '→' in reason:
                        return (1, -item[1])
                    elif '↓' in reason:
                        return (2, -item[1])
                    else:
                        return (3, -item[1])

                for reason, count in sorted(items, key=sort_key):
                    self.log(f"  {reason}: {count}件")

        self.log(f"\n{'─'*50}")
        self.log(f"合計: {total}件")

    def run_sd_adjustment(self):
        self.log(f"\n{'='*40}")
        self.log("SDの調整を開始")

        df = pd.read_excel(self.full_path, sheet_name=0)
        self.log(f"キャンペーン: {len(df)}行")

        p = {k: v.get() for k, v in self.params.items()}
        col = SD_COL

        df_sorted = df.copy()
        df_sorted['_sales'] = df_sorted.iloc[:, col['SALES']].apply(lambda x: self.safe_float(x, 0))
        top_n_idx = set(df_sorted.sort_values('_sales', ascending=False).head(p['top_n']).index)

        new_bids, original_bids, reasons = [], [], []
        stats = {}

        for idx, row in df.iterrows():
            current = self.safe_float(row.iloc[col['BID']], 0)
            clicks = self.safe_float(row.iloc[col['CLICKS']], 0)
            acos = self.safe_float(row.iloc[col['ACOS']], 0)
            target_acos = self.safe_float(row.iloc[col['TARGET_ACOS']], 0)

            original_bids.append(current)

            reason, new_bid = self._calculate_bid(
                current, clicks, acos, target_acos, True, 3, idx in top_n_idx, p
            )

            new_bids.append(new_bid)
            reasons.append(reason)
            # 理由をそのままキーとして集計
            stats[reason] = stats.get(reason, 0) + 1

        # カテゴリ別に整理してログ出力
        self._log_organized_summary(stats)

        output = self.full_path.replace('.xlsx', '_adjusted.xlsx')
        wb = load_workbook(self.full_path)
        ws = wb.worksheets[0]

        out_col = 49  # 許容ACOS(AV列=48列目)の右側、AW列
        ws.cell(row=1, column=out_col, value="元の入札額")
        ws.cell(row=1, column=out_col+1, value="新入札額")
        ws.cell(row=1, column=out_col+2, value="理由")

        for i, (o, n, r) in enumerate(zip(original_bids, new_bids, reasons)):
            ws.cell(row=i+2, column=out_col, value=o)
            ws.cell(row=i+2, column=out_col+1, value=n)
            ws.cell(row=i+2, column=out_col+2, value=r)

        wb.save(output)
        self.log(f"\n完了: {os.path.basename(output)}")
        messagebox.showinfo("完了", f"処理が完了しました\n\n{output}")


if __name__ == "__main__":
    root = TkinterDnD.Tk()
    app = AmazonBidAdjusterApp(root)
    root.mainloop()
